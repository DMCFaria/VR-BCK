import os
import re
import logging
from decimal import Decimal, InvalidOperation
from datetime import datetime
import openpyxl

logger = logging.getLogger(__name__)


# Fallback para produto Refeição-like sem código próprio identificado.
CODIGO_PRODUTO_PADRAO = '31'

# Mapeamento dos headers de produto encontrados nas planilhas para o
# código do produto usado nos registros 50/60 do TXT de compra.
# Códigos conforme a tabela oficial da VR (PRODUTOS - VR.xlsx, ago/2026).
COLUNAS_PRODUTO = {
    # ===== Sem prefixo "VR " (VR-exemplo.xlsm / Template_VR_OLD.xlsx) =====
    'Refeição': '31',
    'Multi Refeição': '244',
    'Alimentação': '27',
    'Multi Alimentação': '212',
    'Auto': '28',
    'Mobilidade': '262',
    'VR Mobilidade': '262',
    'Multi Mobilidade': '262',
    'VR Multi Mobilidade': '262',
    'Cesta': '201',
    'Boas Festas': '202',
    'Multi Boas Festas': '217',
    'Auxílio Alimentação': '204',
    'Multi Auxílio Alimentação': '211',
    'Auxílio Refeição': '243',
    'Multi Auxílio Refeição': '245',
    'Multibenefício': '207',
    'Multibenefícios': '207',
    'Auxílio VR+VA': '209',
    'Multi Auxílio VR+VA': '213',
    'Multi - Home Office': '58',
    'Multi Home office': '58',
    'Refeição Adicional': '242',
    'Auto Manutenção': '261',
    # ===== Com prefixo "VR " (Template_VR.xlsm) =====
    'VR Multi Home office': '58',
    'VR Refeição': '31',
    'VR Alimentação': '27',
    'VR Auto': '28',
    'VR Alimentação Cesta': '201',
    'VR Boas Festas': '202',
    'VR Multi Boas Festas': '217',
    'VR Auxílio Alimentação': '204',
    'VR Auxílio Refeição': '243',
    'VR Multibenefícios': '207',
    'VR+VA': '209',
    'VR Multi Refeição': '244',
    'VR Multi Alimentação': '212',
    'VR Multi Alimentação Valor do crédito': '212',
    'VR Multi Refeição Auxílio': '245',
    'VR Multi Alimentação Auxílio': '211',
    'VR Multi VR+VA': '213',
    'VR Multi - Home Office': '58',
    'VR Refeição Adicional': '242',
    'VR Auto Manutenção': '261',
}

# Mapeamento dos headers de produto para o TIPO normalizado.
# Esse tipo é o valor que deve aparecer em todos os documentos exportados.
MAPEAMENTO_PRODUTO_TIPO = {
    'Refeição': 'Refeição',
    'Multi Refeição': 'Multi - Refeição',
    'Alimentação': 'Alimentação',
    'Multi Alimentação': 'Multi - Alimentação',
    'Auto': 'Auto',
    'Mobilidade': 'Multi - Mobilidade',
    'VR Mobilidade': 'Multi - Mobilidade',
    'Multi Mobilidade': 'Multi - Mobilidade',
    'VR Multi Mobilidade': 'Multi - Mobilidade',
    'Cesta': 'Boas Festas',
    'Boas Festas': 'Boas Festas',
    'Auxílio Alimentação': 'Alimentação',
    'Multi Auxílio Alimentação': 'Multi - Alimentação',
    'Auxílio Refeição': 'Refeição',
    'Multi Auxílio Refeição': 'Multi - Refeição',
    'Multibenefício': 'Multi - VR+VA',
    'Multibenefícios': 'Multi - VR+VA',
    'Auxílio VR+VA': 'Multi - VR+VA',
    'Multi Auxílio VR+VA': 'Multi - VR+VA',
    'Multi - Home Office': 'Multi - Home Office',
    'Multi Home office': 'Multi - Home Office',
    'VR Multi Home office': 'Multi - Home Office',
    'VR Refeição': 'Refeição',
    'VR Alimentação': 'Alimentação',
    'VR Auto': 'Auto',
    'VR Alimentação Cesta': 'Boas Festas',
    'VR Boas Festas': 'Boas Festas',
    'VR Auxílio Alimentação': 'Alimentação',
    'VR Auxílio Refeição': 'Refeição',
    'VR Multibenefícios': 'Multi - VR+VA',
    'VR+VA': 'Multi - VR+VA',
    'VR Multi Refeição': 'Multi - Refeição',
    'VR Multi Alimentação': 'Multi - Alimentação',
    'VR Multi Alimentação Valor do crédito': 'Multi - Alimentação',
    'VR Multi Refeição Auxílio': 'Multi - Refeição',
    'VR Multi Alimentação Auxílio': 'Multi - Alimentação',
    'VR Multi VR+VA': 'Multi - VR+VA',
    'VR Multi - Home Office': 'Multi - Home Office',
    # Produtos de baixa aderência que chegam por cabeçalho
    'Multi Boas Festas': 'Boas Festas',
    'VR Multi Boas Festas': 'Boas Festas',
    'Refeição Adicional': 'Refeição',
    'VR Refeição Adicional': 'Refeição',
    'Auto Manutenção': 'Auto',
    'VR Auto Manutenção': 'Auto',
}

# Headers de produtos que devem ser reconhecidos mas rejeitados com erro claro.
# Útil para produtos não contratados/não suportados no momento.
PRODUTOS_REJEITADOS = {
    'Cultura': 'Produto "Cultura" não é permitido. Remova a coluna ou entre em contato com o suporte.',
    'VR Cultura': 'Produto "Cultura" não é permitido. Remova a coluna ou entre em contato com o suporte.',
    'Multi Premiação': 'Produto "Multi Premiação" não é permitido. Use "Multi - Home Office" ou remova a coluna.',
    'VR Multi Premiação': 'Produto "Multi Premiação" não é permitido. Use "VR Multi - Home Office" ou remova a coluna.',
}

# Mapeamento POSICIONAL de colunas (PRIORITÁRIO).
# Coluna J=10 até Z=26, na ordem exata do template VR padrão.
# Se a coluna estiver neste dict, o produto é determinado pela posição,
# INDEPENDENTE do header da coluna.
# Códigos conforme a tabela oficial da VR (PRODUTOS - VR.xlsx, ago/2026):
# cada produto tem código próprio. O mapeamento antigo achatava tudo em
# {27, 28, 201, 202, 204, 207}, misturando produtos distintos no mesmo
# código e corrompendo nome/tipo nas telas e na planilha de faturamento.
COLUNAS_POSICAO = {
    10: {'nome': 'VR Refeição',               'codigo': '31',  'tipo': 'Refeição'},
    11: {'nome': 'VR Alimentação',             'codigo': '27',  'tipo': 'Alimentação'},
    12: {'nome': 'VR Auto',                    'codigo': '28',  'tipo': 'Auto'},
    13: {'nome': 'VR Cultura',                 'codigo': '30',  'tipo': None, 'rejeitado': True},
    14: {'nome': 'VR Alimentação Cesta',       'codigo': '201', 'tipo': 'Boas Festas'},
    15: {'nome': 'VR Boas Festas',             'codigo': '202', 'tipo': 'Boas Festas'},
    16: {'nome': 'VR Auxílio Alimentação',     'codigo': '204', 'tipo': 'Alimentação'},
    17: {'nome': 'VR Auxílio Refeição',        'codigo': '243', 'tipo': 'Refeição'},
    18: {'nome': 'VR Multibenefícios',         'codigo': '207', 'tipo': 'Multi - VR+VA'},
    19: {'nome': 'VR+VA',                      'codigo': '209', 'tipo': 'Multi - VR+VA'},
    20: {'nome': 'VR Multi Refeição',          'codigo': '244', 'tipo': 'Multi - Refeição'},
    21: {'nome': 'VR Multi Alimentação',       'codigo': '212', 'tipo': 'Multi - Alimentação'},
    22: {'nome': 'VR Multi Refeição Auxílio',  'codigo': '245', 'tipo': 'Multi - Refeição'},
    23: {'nome': 'VR Multi Alimentação Auxílio','codigo': '211', 'tipo': 'Multi - Alimentação'},
    24: {'nome': 'VR Multi VR+VA',             'codigo': '213', 'tipo': 'Multi - VR+VA'},
    25: {'nome': 'VR Multi Home office',       'codigo': '58',  'tipo': 'Multi - Home Office'},
    26: {'nome': 'VR Multi Mobilidade',        'codigo': '262', 'tipo': 'Multi - Mobilidade'},
}

# Sheets 60/99/etc são para geração do TXT via VBA - não precisam ser lidas
# Focus: Local de Entrega, Beneficiario, Sumario


def _safe_str(val, default=''):
    if val is None:
        return default
    s = str(val).strip()
    if s.lower() in ('none', 'nan', ''):
        return default
    return s


def _parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S', '%d%m%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except:
            pass
    return None


def _normalizar_cnpj(val):
    """Remove tudo que não for dígito e preenche com zeros à esquerda até 14."""
    if val is None:
        return ''
    return re.sub(r'\D', '', str(val)).zfill(14)[:14]


def _remover_acentos(texto):
    """Remove acentos e caracteres especiais, convertendo para ASCII."""
    if not texto:
        return texto
    import unicodedata
    nfkd = unicodedata.normalize('NFKD', texto)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def _endereco_normalizado(local):
    """
    Tupla de endereço normalizada (sem acentos, casefold) para comparar
    locais de entrega entre si. Usada na detecção de cartão admin.
    """
    def norm(campo):
        return _remover_acentos(_safe_str(local.get(campo))).casefold().strip()
    return tuple(norm(c) for c in ('rua', 'numero', 'bairro', 'cidade', 'cep'))


def _normalizar_header(val):
    """Extrai e normaliza o texto do header para comparação (primeira linha, sem espaços extras)."""
    if val is None:
        return ''
    h = _safe_str(val).strip().split('\n')[0].strip()
    return h


def validar_dimensoes_planilha(file_path, max_abas=50, max_linhas_beneficiario=10000, max_colunas=100):
    """
    Valida dimensões da planilha em modo read_only (baixo consumo de memória)
    antes de processar. Evita SIGKILL por excesso de RAM.

    Retorna dict com 'ok' (bool) e 'erro' (str, se aplicável).
    """
    # Planilha protegida por senha (ou .xls 97-2003 renomeado): o Excel
    # embrulha em contêiner OLE2 e o openpyxl falha com o críptico
    # "File is not a zip file". Detectar pela assinatura e explicar.
    try:
        with open(file_path, 'rb') as f:
            assinatura = f.read(8)
        if assinatura == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
            return {
                "ok": False,
                "erro": (
                    "A planilha está protegida por senha ou no formato antigo "
                    "Excel 97-2003. Remova a senha de abertura (Arquivo → Informações "
                    "→ Proteger Pasta de Trabalho) ou salve como .xlsx/.xlsm e envie novamente."
                ),
            }
    except Exception:
        pass

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
    except Exception as e:
        return {"ok": False, "erro": f"Não foi possível abrir a planilha: {str(e)}"}

    try:
        num_abas = len(wb.sheetnames)
        if num_abas > max_abas:
            return {
                "ok": False,
                "erro": (
                    f"Planilha contém {num_abas} abas. "
                    f"Limite máximo: {max_abas}. "
                    f"Remova abas desnecessárias (só são necessárias: Sumario, Local de Entrega, Beneficiario)."
                ),
            }

        nome_beneficiario = None
        for nome in wb.sheetnames:
            if nome.lower().replace('í', 'i').replace('é', 'e') in ('beneficiario', 'beneficiários'):
                nome_beneficiario = nome
                break

        if nome_beneficiario:
            ws = wb[nome_beneficiario]
            ultima_linha_com_dados = 0
            ultima_coluna = 0
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                has_data = any(v is not None and str(v).strip() != '' for v in row)
                if has_data:
                    ultima_linha_com_dados = row_idx
                    if row_idx > max_linhas_beneficiario:
                        return {
                            "ok": False,
                            "erro": (
                                f"Aba '{nome_beneficiario}' tem mais de {max_linhas_beneficiario} linhas "
                                f"com dados. Reduza a planilha ou entre em contato com o suporte."
                            ),
                        }
                    for col_idx, val in enumerate(row, start=1):
                        if val is not None and str(val).strip() != '':
                            if col_idx > max_colunas:
                                return {
                                    "ok": False,
                                    "erro": (
                                        f"Aba '{nome_beneficiario}' contém dados na coluna {col_idx} "
                                        f"(limite: {max_colunas}). Verifique se há colunas formatadas "
                                        f"além do necessário ou entre em contato com o suporte."
                                    ),
                                }
                            ultima_coluna = col_idx

            logger.info(
                f"[VALIDACAO] Planilha OK: {num_abas} abas, "
                f"'{nome_beneficiario}'={ultima_linha_com_dados} linhas x {ultima_coluna} colunas"
            )
        else:
            logger.warning("[VALIDACAO] Aba 'Beneficiario' não encontrada, pulando validação de dimensões")

        return {"ok": True}

    finally:
        wb.close()


def parse_fut_template(file_path, file_upload_id, valor_max_beneficio=None, administradora_cnpj=None):
    """
    Lê o template VR em modo read_only (o template traz ~40 abas de VBA e o
    modo normal carrega todas — 11s+ num arquivo pequeno, estourando o timeout
    de 30s do gunicorn em produção). Alguns arquivos têm o metadado de
    dimensões quebrado e voltam VAZIOS no read_only: nesses casos, sem nenhum
    erro nem condomínio no resultado, reprocessa no modo completo.
    """
    resultado = _parse_fut_template(
        file_path, file_upload_id, valor_max_beneficio, administradora_cnpj,
        read_only=True,
    )

    # Sem nenhum condomínio nem linha processada = leitura suspeita (arquivos
    # com metadado quebrado voltam vazios e ainda geram erros artificiais tipo
    # "aba sem locais"). O reparse completo confirma: se a planilha for vazia
    # de verdade, o modo completo devolve o mesmo erro legítimo.
    vazio_suspeito = (
        not resultado.get('condominios')
        and not resultado.get('linhas_com_erro')
    )
    if vazio_suspeito:
        logger.warning(
            "[READER] Parse read_only voltou vazio — refazendo em modo completo "
            "(possível metadado de dimensões quebrado no arquivo)."
        )
        resultado = _parse_fut_template(
            file_path, file_upload_id, valor_max_beneficio, administradora_cnpj,
            read_only=False,
        )

    return resultado


def _parse_fut_template(file_path, file_upload_id, valor_max_beneficio=None, administradora_cnpj=None, read_only=True):
    if valor_max_beneficio is None:
        valor_max_beneficio = Decimal('9999.99')

    result = {
        "file_upload_id": file_upload_id,
        "cartao_admin": None,
        "administradora_cnpj": None,
        "condominios": [],
        "errors": [],
        "linhas_com_erro": [],
        "erros_condominios": [],
        "avisos": [],
        "summary": {
            "total_condominios": 0,
            "total_funcionarios": 0,
            "total_movimentacoes": 0,
            "valor_total_beneficios": Decimal('0.00'),
            "data_competencia_arquivo": None,
            "primeiro_cnpj_processado": "N/A",
        },
    }

    file_ext = os.path.splitext(file_path)[1].lower()

    if file_ext in ('.csv', '.xls'):
        result['errors'].append(
            f"Formato '{file_ext}' não suportado. Use arquivos .xlsx ou .xlsm no layout VR."
        )
        return result

    try:
        # read_only: o template VR traz ~40 abas (VBA/registros 60/99) e o modo
        # normal carrega todas — 11s+ num arquivo de 269 linhas, estourando o
        # timeout de 30s do gunicorn em produção (upload preso em PENDING +
        # erro 500). Só usamos iter_rows nas 3 abas de interesse, que o modo
        # read_only atende, como a validar_dimensoes_planilha já faz.
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=read_only)
    except Exception as e:
        result['errors'].append(f"Erro ao abrir planilha: {str(e)}")
        return result

    # ========================
    # 1. LER SUMARIO
    # ========================
    ws_sum = wb['Sumario']
    data_disponivel = ''
    cnpj_sumario = ''
    for row in ws_sum.iter_rows(min_row=1, max_row=6, values_only=True):
        if not row:
            continue
        primeiro_valor = _safe_str(row[0] if len(row) > 0 else '')
        if 'CNPJ' in primeiro_valor.upper() or 'CÓDIGO DO CLIENTE' in primeiro_valor.upper():
            for cell in row[1:]:
                val = _safe_str(cell)
                if val:
                    cnpj_sumario = _normalizar_cnpj(val)
                    if cnpj_sumario:
                        break
            continue
        # data disponível costuma estar na linha 6
        for cell in row:
            val = _safe_str(cell)
            if val:
                parsed = _parse_date(val)
                if parsed:
                    data_disponivel = val
                    break

    result['summary']['data_competencia_arquivo'] = _parse_date(data_disponivel)

    admin_cnpj_normalizado = _normalizar_cnpj(administradora_cnpj or cnpj_sumario)
    result['administradora_cnpj'] = admin_cnpj_normalizado

    # ========================
    # 2. LER LOCAIS DE ENTREGA (condominios)
    # ========================
    ws_locais = wb['Local de Entrega']
    locais = {}  # codigo_local -> dados
    locais_raw = []  # lista de codigos na ordem da planilha

    for row in ws_locais.iter_rows(min_row=2, values_only=True):
        codigo = _safe_str(row[0])
        if not codigo:
            continue

        locais_raw.append(codigo)
        locais[codigo] = {
            "nome": _safe_str(row[1]) if len(row) > 1 else codigo,
            "cnpj": codigo,
            "valor_condo": Decimal('0.00'),
            "rua": _safe_str(row[3]) if len(row) > 3 else '',
            "numero": _safe_str(row[4]) if len(row) > 4 else '',
            "complemento": _safe_str(row[5]) if len(row) > 5 else '',
            "bairro": _safe_str(row[6]) if len(row) > 6 else '',
            "cidade": _safe_str(row[7]) if len(row) > 7 else '',
            "estado": _safe_str(row[8]) if len(row) > 8 else '',
            "cep": _safe_str(row[9]) if len(row) > 9 else '',
            "funcionarios": {},
        }

    # Determina se é modo "cartão admin" baseado apenas na aba Local de Entrega.
    # Regra acordada: se houver apenas 1 local de entrega => cartao_admin=True
    # (a administradora centraliza a entrega). Se houver 0 locais => erro.
    #
    # Com >1 local, normalmente são condomínios com endereços próprios
    # (cartao_admin=False). Mas existe um formato híbrido em que a planilha
    # lista um local por condomínio e repete o endereço da ADMINISTRADORA em
    # todos eles — nesse caso o endereço não é do condomínio e a entrega
    # também é centralizada, então é cartão admin do mesmo jeito. Endereços
    # todos vazios NÃO contam: seguem como condomínios normais, pois o
    # faturamento já consulta o CNPJ quando o endereço está vazio.
    cartao_admin = False
    if not locais_raw:
        result['errors'].append(
            "Aba 'Local de Entrega' não contém nenhum local de entrega cadastrado."
        )
    elif len(locais_raw) == 1:
        cartao_admin = True
    else:
        enderecos = {_endereco_normalizado(locais[codigo]) for codigo in locais_raw}
        if len(enderecos) == 1:
            (endereco_unico,) = enderecos
            rua = endereco_unico[0]
            cartao_admin = bool(rua)

    result['cartao_admin'] = cartao_admin

    # ========================
    # 3. LER BENEFICIARIO
    # ========================
    ws_ben = wb['Beneficiario']

    MAP_BEN = {
        'cpf': 0,
        'codigo_local': 1,
        'centro_custo': 2,
        'matricula': 3,
        'nome_completo': 4,
        'nome_impressao': 5,
        'data_nascimento': 6,
        'sexo': 7,
        'faixa_salarial': 8,
    }

    # ============================
    # 3.1. MAPEAMENTO DE COLUNAS DE PRODUTO
    # ============================
    # PRIORIDADE 1: Mapeamento posicional (COLUNAS_POSICAO, colunas J=10 a Z=26)
    # PRIORIDADE 2: Reconhecimento por header (fallback para templates não padrão)
    col_produtos = {}   # col_idx -> nome_produto (string)
    col_rejeitados = {} # col_idx -> nome_produto rejeitado

    ben_rows = list(ws_ben.iter_rows(min_row=1, max_row=5, values_only=True))

    def _match_produto_header(val):
        """Fallback: retorna o nome do produto se o header for reconhecido."""
        if val is None:
            return None
        h = _normalizar_header(val)
        if not h:
            return None
        # Match exato
        if h in COLUNAS_PRODUTO:
            return h
        # Match por normalização (acentos + case)
        h_norm = _remover_acentos(h).lower()
        for nome in COLUNAS_PRODUTO:
            if _remover_acentos(nome).lower() == h_norm:
                return nome
        # Match parcial (substring) - último recurso
        for nome in COLUNAS_PRODUTO:
            if _remover_acentos(nome).lower() in h_norm:
                return nome
        return None

    def _match_produto_rejeitado(val):
        """Retorna o nome do produto se for um header de produto rejeitado."""
        if val is None:
            return None
        h = _normalizar_header(val)
        if not h:
            return None
        if h in PRODUTOS_REJEITADOS:
            return h
        h_norm = _remover_acentos(h).lower()
        for nome in PRODUTOS_REJEITADOS:
            if _remover_acentos(nome).lower() in h_norm:
                return nome
        return None

    # Detectar linha de header (usa apenas reconhecimento por header, NÃO posicional)
    # O mapeamento posicional é usado DEPOIS para atribuir produtos às colunas.
    best_row = None
    best_row_num = 2
    best_count = 0
    for try_idx, try_row in enumerate(ben_rows):
        if not try_row:
            continue
        matches = [_match_produto_header(v) for v in try_row]
        count = sum(1 for m in matches if m)
        # Só considera se tiver pelo menos 2 produtos (evita falsos positivos)
        if count >= 2 and count > best_count:
            best_row = try_row
            best_row_num = try_idx + 1
            best_count = count

    if not best_row and ben_rows:
        # Fallback: usa a primeira linha que tiver algum match
        for try_idx, try_row in enumerate(ben_rows):
            if any(_match_produto_header(v) for v in try_row):
                best_row = try_row
                best_row_num = try_idx + 1
                break
        if not best_row:
            best_row = ben_rows[1] if len(ben_rows) > 1 else ben_rows[0]
            best_row_num = 2

    header_row = best_row
    header_row_num = best_row_num

    # Mapear colunas: POSIÇÃO (prioritário) > HEADER (fallback)
    if header_row:
        for col_idx, val in enumerate(header_row, start=1):
            # PRIORIDADE 1: Mapeamento posicional
            if col_idx in COLUNAS_POSICAO:
                info = COLUNAS_POSICAO[col_idx]
                if info.get('rejeitado'):
                    col_rejeitados[col_idx] = info['nome']
                else:
                    col_produtos[col_idx] = info['nome']
            else:
                # PRIORIDADE 2: Fallback por header
                produto = _match_produto_header(val)
                if produto:
                    col_produtos[col_idx] = produto
                else:
                    rejeitado = _match_produto_rejeitado(val)
                    if rejeitado:
                        col_rejeitados[col_idx] = rejeitado

    data_start_row = header_row_num + 1

    # Se a primeira linha de dados não tem CPF/nome/CNPJ, pular para a próxima
    # (_templates com linha de descrição entre header e dados)
    probe_rows = list(ws_ben.iter_rows(min_row=data_start_row, max_row=data_start_row + 2, values_only=True))
    for probe_idx, probe_row in enumerate(probe_rows):
        if not probe_row:
            continue
        cpf_probe = _safe_str(probe_row[0] if len(probe_row) > 0 else '')
        nome_probe = _safe_str(probe_row[4] if len(probe_row) > 4 else '')
        cnpj_probe = _safe_str(probe_row[1] if len(probe_row) > 1 else '')
        cpf_digits = re.sub(r'\D', '', cpf_probe)
        if cpf_digits and len(cpf_digits) == 11 and nome_probe and cnpj_probe:
            data_start_row = header_row_num + 1 + probe_idx
            break
    else:
        # Se nenhuma linha de dados válida foi encontrada nas 3 tentativas,
        # manter data_start_row original (será tratado como erro abaixo)
        pass

    # ============================
    # 3.2. VALIDAÇÃO PRÉVIA: COLUNAS SEM MAPEAMENTO
    # ============================
    # Verificar se há colunas entre 10-26 que não foram mapeadas
    # (nem por posição, nem por header) e que possuem valores > 0 nos dados.
    colunas_nao_mapeadas = set()
    data_rows = list(ws_ben.iter_rows(min_row=data_start_row, values_only=True))
    for row in data_rows:
        if not row:
            continue
        for col_idx in range(10, min(27, len(row) + 1)):
            if col_idx in col_produtos or col_idx in col_rejeitados:
                continue
            raw_val = row[col_idx - 1] if len(row) > col_idx - 1 else None
            if raw_val is None:
                continue
            try:
                valor = Decimal(str(raw_val).replace(',', '.'))
                if valor > 0:
                    colunas_nao_mapeadas.add(col_idx)
            except:
                pass

    if colunas_nao_mapeadas:
        for col_idx in sorted(colunas_nao_mapeadas):
            header_text = _normalizar_header(
                header_row[col_idx - 1] if header_row and len(header_row) > col_idx - 1 else None
            )
            result['errors'].append(
                f"Coluna {col_idx} (\"{header_text or 'sem header'}\") contém valores "
                f"mas não está mapeada nem como produto nem como rejeitado. "
                f"Verifique o template ou adicione o mapeamento."
            )
    for row_num, row in enumerate(data_rows, start=data_start_row):
        if not row:
            continue

        # Pular linhas completamente vazias (memória do Excel)
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        cpf_raw = _safe_str(row[0] if len(row) > 0 else '')
        codigo_local = _safe_str(row[1] if len(row) > 1 else '')
        matricula = _safe_str(row[3] if len(row) > 3 else '')
        nome = _safe_str(row[4] if len(row) > 4 else '')
        data_nasc_raw = row[6] if len(row) > 6 else None

        erros_linha_atual = []

        # 1. Validar CPF
        cpf = ""
        if not cpf_raw:
            erros_linha_atual.append("CPF do beneficiário ausente")
        else:
            cpf = re.sub(r'\D', '', cpf_raw)
            if len(cpf) < 11:
                cpf = cpf.zfill(11)
            if len(cpf) != 11:
                erros_linha_atual.append(f"CPF inválido (tamanho incorreto: {len(cpf)} dígitos)")

        # 2. Validar Nome
        if not nome:
            erros_linha_atual.append("Nome do beneficiário ausente")

        # 3. Validar Código Local de Entrega
        if not codigo_local:
            erros_linha_atual.append("Código local de entrega (CNPJ) ausente")

        # 5. Validar Data de Nascimento
        data_nasc = _parse_date(data_nasc_raw)
        if not data_nasc:
            erros_linha_atual.append("Data de nascimento ausente ou inválida")

        # 6. Validar se há pelo menos um benefício/produto com valor > 0
        tem_beneficio = False
        for col_idx, nome_produto in col_produtos.items():
            raw_val = row[col_idx - 1] if len(row) > col_idx - 1 else None
            if raw_val is not None:
                try:
                    valor = Decimal(str(raw_val).replace(',', '.'))
                    if valor > 0:
                        tem_beneficio = True
                        break
                except:
                    pass

        if not tem_beneficio:
            erros_linha_atual.append("Nenhum benefício preenchido com valor maior que zero")

        # 7. Validar produtos rejeitados (ex: Cultura, Premiação)
        for col_idx, nome_rejeitado in col_rejeitados.items():
            raw_val = row[col_idx - 1] if len(row) > col_idx - 1 else None
            if raw_val is not None:
                try:
                    valor = Decimal(str(raw_val).replace(',', '.'))
                    if valor > 0:
                        erros_linha_atual.append(PRODUTOS_REJEITADOS[nome_rejeitado])
                except:
                    pass

        if erros_linha_atual:
            result['linhas_com_erro'].append({
                "tipo_erro": "INFORMACAO_AUSENTE",
                "linha": row_num,
                "dados": {
                    "cpf": cpf_raw,
                    "nome": nome,
                    "codigo_local": codigo_local,
                    "matricula": matricula,
                    "data_nascimento": _safe_str(data_nasc_raw),
                },
                "erros": erros_linha_atual
            })
        else:
            if codigo_local not in locais:
                # No modo cartão admin os condomínios vêm apenas da aba Beneficiário.
                # A coluna "Matrícula" costuma trazer o nome descritivo do condomínio,
                # então usamos ela como nome quando disponível.
                nome_local = matricula.strip() if matricula and str(matricula).strip() else codigo_local
                locais[codigo_local] = {
                    "nome": nome_local,
                    "cnpj": codigo_local,
                    "valor_condo": Decimal('0.00'),
                    "rua": '',
                    "numero": '',
                    "complemento": '',
                    "bairro": '',
                    "cidade": '',
                    "estado": '',
                    "cep": '',
                    "funcionarios": {},
                    "nao_cadastrado_local_entrega": True,
                }

            func_key = f"{codigo_local}_{cpf}"
            if func_key not in locais[codigo_local]["funcionarios"]:
                locais[codigo_local]["funcionarios"][func_key] = {
                    "nome": nome.upper() if nome else '',
                    "cpf": cpf,
                    "matricula": matricula,
                    "departamento": '',
                    "funcao": '',
                    "data_nascimento": data_nasc,
                    "cep": None,
                    "endereco_rua": None,
                    "endereco_numero": None,
                    "endereco_complemento": None,
                    "endereco_bairro": None,
                    "valor_bene": Decimal('0.00'),
                    "movimentacoes": [],
                    "linha": row_num,
                    "linhas_somadas": [],
                    "conflito_cpf": False,
                }
            else:
                # CPF repetido no mesmo condomínio: só é a mesma pessoa se nome
                # e nascimento baterem. Divergência = CPF errado para pessoas
                # distintas — somar aqui creditaria duas pessoas num CPF só.
                existente = locais[codigo_local]["funcionarios"][func_key]
                nome_atual = (nome or '').strip().upper()
                nome_anterior = (existente["nome"] or '').strip().upper()
                nascimento_diverge = bool(
                    data_nasc and existente["data_nascimento"]
                    and data_nasc != existente["data_nascimento"]
                )

                if nome_atual != nome_anterior or nascimento_diverge:
                    existente["conflito_cpf"] = True
                    result['linhas_com_erro'].append({
                        "tipo_erro": "CPF_DUPLICADO_DIVERGENTE",
                        "linha": row_num,
                        "dados": {
                            "cpf": cpf_raw,
                            "nome": nome,
                            "codigo_local": codigo_local,
                            "matricula": matricula,
                            "data_nascimento": _safe_str(data_nasc_raw),
                        },
                        "erros": [
                            (
                                f"CPF duplicado com dados divergentes: já usado na linha "
                                f"{existente['linha']} por '{existente['nome']}'"
                                + (f" (nasc. {existente['data_nascimento']})" if existente['data_nascimento'] else "")
                                + " — mesmo CPF para pessoas distintas; corrija o CPF na planilha."
                            )
                        ],
                    })
                    continue

                existente["linhas_somadas"].append(row_num)

            func = locais[codigo_local]["funcionarios"][func_key]

            for col_idx, nome_produto in col_produtos.items():
                raw_val = row[col_idx - 1] if len(row) > col_idx - 1 else None
                if raw_val is None:
                    continue
                try:
                    valor = Decimal(str(raw_val).replace(',', '.'))
                except:
                    continue
                if valor == 0:
                    continue

                # Obter código/tipo: prioriza COLUNAS_POSICAO, fallback para dicts
                if col_idx in COLUNAS_POSICAO:
                    info = COLUNAS_POSICAO[col_idx]
                    codigo_produto = info['codigo']
                    tipo_produto = info['tipo']
                else:
                    codigo_produto = COLUNAS_PRODUTO.get(nome_produto, '')
                    tipo_produto = MAPEAMENTO_PRODUTO_TIPO.get(nome_produto, nome_produto)

                # DEDUPLICAÇÃO: se já existe movimentação para este mesmo produto,
                # somar o valor ao invés de criar uma nova entrada.
                existing_mov = next(
                    (m for m in func["movimentacoes"] if m["produto"] == nome_produto),
                    None,
                )
                if existing_mov:
                    existing_mov["valor"] += valor
                else:
                    func["movimentacoes"].append({
                        "produto": nome_produto,
                        "tipo": tipo_produto,
                        "codigo_produto": codigo_produto,
                        "valor": valor,
                    })
                func["valor_bene"] += valor
                locais[codigo_local]["valor_condo"] += valor
                result['summary']['valor_total_beneficios'] += valor

    wb.close()

    # ========================
    # 4. MONTAR CONDOMINIOS
    # ========================
    erros_condominios = []
    for codigo, local in locais.items():
        lista_func = []
        for fkey, func in local["funcionarios"].items():
            if not func["movimentacoes"]:
                continue

            # Conflito de CPF (mesmo CPF, pessoa diferente): a 1ª ocorrência
            # também sai do lote — não dá para saber qual das duas está com o
            # CPF certo. Estorna os valores já acumulados e registra o erro.
            if func.get("conflito_cpf"):
                local["valor_condo"] -= func["valor_bene"]
                result['summary']['valor_total_beneficios'] -= func["valor_bene"]
                result['linhas_com_erro'].append({
                    "tipo_erro": "CPF_DUPLICADO_DIVERGENTE",
                    "linha": func.get("linha"),
                    "dados": {
                        "cpf": func["cpf"],
                        "nome": func["nome"],
                        "codigo_local": codigo,
                        "matricula": func["matricula"],
                        "data_nascimento": _safe_str(func["data_nascimento"]),
                    },
                    "erros": [
                        "CPF duplicado com dados divergentes: o mesmo CPF aparece em "
                        "outra linha com nome/nascimento diferente — mesmo CPF para "
                        "pessoas distintas; corrija o CPF na planilha."
                    ],
                })
                continue

            if func.get("linhas_somadas"):
                result['avisos'].append({
                    "tipo": "CPF_SOMADO",
                    "cpf": func["cpf"],
                    "nome": func["nome"],
                    "condominio": local.get("nome") or codigo,
                    "linhas": [func.get("linha")] + func["linhas_somadas"],
                    "valor_total": func["valor_bene"],
                })

            lista_func.append({
                "nome": func["nome"],
                "cpf": func["cpf"],
                "matricula": func["matricula"],
                "departamento": func["departamento"],
                "funcao": func["funcao"],
                "data_nascimento": func["data_nascimento"],
                "cep": func["cep"],
                "endereco_rua": func["endereco_rua"],
                "endereco_numero": func["endereco_numero"],
                "endereco_complemento": func["endereco_complemento"],
                "endereco_bairro": func["endereco_bairro"],
                "valor_bene": func["valor_bene"],
                "movimentacoes": func["movimentacoes"],
            })

        if lista_func:
            missing_infos = []
            if not cartao_admin and local.get("nao_cadastrado_local_entrega"):
                # No modo normal (condomínios) exigimos que o local esteja na aba Local de Entrega
                missing_infos.append("Não cadastrado na aba 'Local de Entrega'")
            else:
                if not local.get("nome") or local["nome"].startswith("Local: "):
                    missing_infos.append("Nome do condomínio ausente ou inválido")
                if not local.get("cnpj"):
                    missing_infos.append("CNPJ do condomínio ausente")
                # No modo cartão admin o endereço será pesquisado automaticamente;
                # ainda assim, se a planilha trouxer endereço, usamos ela como fonte fiel.
                if not cartao_admin:
                    if not local.get("estado"):
                        missing_infos.append("Estado ausente")
                    if not local.get("cep"):
                        missing_infos.append("CEP ausente")

            if missing_infos:
                erros_condominios.append({
                    "cnpj": local["cnpj"],
                    "nome": local["nome"],
                    "erros": missing_infos
                })

            result["condominios"].append({
                "nome": local["nome"],
                "cnpj": local["cnpj"],
                "valor_condo": local["valor_condo"],
                "rua": local.get("rua"),
                "numero": local.get("numero"),
                "complemento": local.get("complemento"),
                "bairro": local.get("bairro"),
                "cidade": local.get("cidade"),
                "estado": local.get("estado"),
                "cep": local.get("cep"),
                "funcionarios": lista_func,
            })
            result['summary']['total_funcionarios'] += len(lista_func)
            result['summary']['total_movimentacoes'] += sum(len(f["movimentacoes"]) for f in lista_func)

    result["erros_condominios"] = erros_condominios
    result['summary']['total_condominios'] = len(result["condominios"])
    if result["condominios"]:
        result['summary']['primeiro_cnpj_processado'] = result["condominios"][0].get("cnpj", "N/A")

    # Se há erros acumulados (erros gerais, linhas com erro ou erros de condomínio),
    # retornamos o payload focado nos erros.
    # Exceção: CPF_DUPLICADO_DIVERGENTE não derruba a planilha inteira — as
    # linhas em conflito já foram excluídas do lote e seguem visíveis em
    # linhas_com_erro para o aviso em tela; o restante do lote continua.
    erros_fatais = [
        l for l in result["linhas_com_erro"]
        if l.get("tipo_erro") != "CPF_DUPLICADO_DIVERGENTE"
    ]
    if result["errors"] or erros_fatais or result["erros_condominios"]:
        mensagem_erro = "Planilha contém informações obrigatórias ausentes ou incorretas."
        if result["errors"]:
            mensagem_erro = "; ".join(result["errors"])
        return {
            "file_upload_id": file_upload_id,
            "cartao_admin": result["cartao_admin"],
            "administradora_cnpj": result["administradora_cnpj"],
            "status": "ERRO",
            "condominios": [],
            "errors": [mensagem_erro],
            "linhas_com_erro": result["linhas_com_erro"],
            "erros_condominios": result["erros_condominios"],
            "summary": {
                "total_condominios": 0,
                "total_funcionarios": 0,
                "total_movimentacoes": 0,
                "valor_total_beneficios": Decimal('0.00'),
                "data_competencia_arquivo": result['summary']['data_competencia_arquivo'],
                "primeiro_cnpj_processado": "N/A",
            }
        }

    return result


def test_parse():
    import json
    import os
    from decimal import Decimal

    file_path = os.path.join(os.path.dirname(__file__), 'VR-exemplo.xlsm')
    if not os.path.exists(file_path):
        file_path = os.path.join(os.path.dirname(__file__), 'Template_VR.xlsm')
    if not os.path.exists(file_path):
        print(f"ERRO: Arquivo não encontrado")
        return

    data = parse_fut_template(
        file_path,
        file_upload_id=999,
        valor_max_beneficio=Decimal('9999.99'),
        administradora_cnpj='35315360000167',
    )

    # resumo
    s = data['summary']
    print(f"cartao_admin: {data.get('cartao_admin')}")
    print(f"administradora_cnpj: {data.get('administradora_cnpj')}")
    print(f"Total condominios: {s['total_condominios']}")
    print(f"Total funcionarios: {s['total_funcionarios']}")
    print(f"Total movimentacoes: {s['total_movimentacoes']}")
    print(f"Valor total beneficios: {s['valor_total_beneficios']}")
    print(f"Data competencia arquivo: {s['data_competencia_arquivo']}")
    print(f"Erros: {len(data['errors'])}")
    print(f"Linhas com erro: {len(data['linhas_com_erro'])}")
    print()

    for c in data['condominios']:
        print(f"Condominio: {c['nome']} (CNPJ: {c['cnpj']})")
        print(f"  Endereco: {c['rua']}, {c['numero']} - {c['bairro']}, {c['cidade']}/{c['estado']} CEP:{c['cep']}")
        print(f"  Funcionarios: {len(c['funcionarios'])}")
        for f in c['funcionarios']:
            print(f"    - {f['nome']} | CPF: {f['cpf']} | Nasc: {f['data_nascimento']} | Mat: {f['matricula']}")
            for m in f['movimentacoes']:
                print(f"        > {m['produto']}: R$ {m['valor']}")
        print()

    print("=== JSON completo ===")
    safe = json.dumps(data, default=str, indent=2, ensure_ascii=False)
    print(safe)


if __name__ == '__main__':
    test_parse()
