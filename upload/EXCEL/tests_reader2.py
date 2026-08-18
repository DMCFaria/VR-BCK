import os
import unittest
from decimal import Decimal

from upload.EXCEL.reader2 import (
    parse_fut_template,
    COLUNAS_POSICAO,
    COLUNAS_PRODUTO,
    MAPEAMENTO_PRODUTO_TIPO,
)


class Reader2HomeOfficeTest(unittest.TestCase):
    """Testes para o parser de planilhas VR."""

    def test_multi_home_office_sem_hifen_eh_reconhecido(self):
        """
        Planilhas que usam 'Multi Home office' (sem hífen) devem reconhecer
        o produto e não retornar erro de 'nenhum benefício preenchido'.
        """
        file_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            'VR - ADM E CONDS.xlsm'
        )
        if not os.path.exists(file_path):
            self.skipTest(f'Arquivo de exemplo não encontrado: {file_path}')

        data = parse_fut_template(
            file_path,
            file_upload_id=1491,
            valor_max_beneficio=Decimal('9999.99'),
            administradora_cnpj='35315360000167',
        )

        self.assertEqual(data.get('status'), None)
        self.assertEqual(len(data.get('linhas_com_erro', [])), 0)
        self.assertEqual(data['summary']['total_funcionarios'], 287)
        self.assertEqual(data['summary']['total_movimentacoes'], 321)

        # Conta movimentações de home office
        home_office_count = 0
        for c in data.get('condominios', []):
            for f in c.get('funcionarios', []):
                for m in f.get('movimentacoes', []):
                    if m.get('tipo') == 'Multi - Home Office':
                        home_office_count += 1

        self.assertEqual(home_office_count, 45)

    def test_produtos_rejeitados_geram_erro(self):
        """
        Colunas de produtos rejeitados (Cultura, Multi Premiação) devem gerar
        erro claro quando possuem valores maiores que zero.
        """
        import openpyxl
        import shutil
        import tempfile

        src = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            'VR - ADM E CONDS.xlsm'
        )
        if not os.path.exists(src):
            self.skipTest(f'Arquivo de exemplo não encontrado: {src}')

        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = os.path.join(tmpdir, 'test_rejeicao.xlsm')
            shutil.copy(src, test_file)

            wb = openpyxl.load_workbook(test_file, data_only=False, keep_vba=True)
            ws = wb['Beneficiario']
            # Coluna Cultura (col 13) na linha 7
            ws.cell(7, 13).value = 100.0
            # Coluna Multi Premiação (col 27) na linha 8
            ws.cell(8, 27).value = 200.0
            wb.save(test_file)

            data = parse_fut_template(
                test_file,
                file_upload_id=1492,
                valor_max_beneficio=Decimal('9999.99'),
                administradora_cnpj='35315360000167',
            )

            self.assertEqual(data.get('status'), 'ERRO')
            erros_por_linha = {
                e['linha']: e['erros']
                for e in data.get('linhas_com_erro', [])
            }
            self.assertIn(7, erros_por_linha)
            self.assertIn(8, erros_por_linha)
            self.assertTrue(
                any('Cultura' in err for err in erros_por_linha[7]),
                f'Esperado erro de Cultura na linha 7: {erros_por_linha[7]}'
            )
            self.assertTrue(
                any('Premiação' in err for err in erros_por_linha[8]),
                f'Esperado erro de Premiação na linha 8: {erros_por_linha[8]}'
            )


class TestColunasPosicao(unittest.TestCase):
    """Testes para o mapeamento posicional de colunas."""

    def test_colunas_posicao_cobre_colunas_j_a_z(self):
        """COLUNAS_POSICAO deve mapear colunas 10 (J) a 26 (Z)."""
        for col_idx in range(10, 27):
            self.assertIn(col_idx, COLUNAS_POSICAO,
                          f'Coluna {col_idx} não está em COLUNAS_POSICAO')

    def test_colunas_posicao_tipos_validos(self):
        """Todos os tipos em COLUNAS_POSICAO devem ser válidos no Produto."""
        tipos_validos = {display for _, display in [
            ('ALIMENTACAO', 'Alimentação'),
            ('AUTO', 'Auto'),
            ('REFEICAO', 'Refeição'),
            ('MULTI_HOME_OFFICE', 'Multi - Home Office'),
            ('BOAS_FESTAS', 'Boas Festas'),
            ('MULTI_ALIMENTACAO', 'Multi - Alimentação'),
            ('MULTI_VR_VA', 'Multi - VR+VA'),
            ('MULTI_REFEICAO', 'Multi - Refeição'),
            ('MULTI_MOBILIDADE', 'Multi - Mobilidade'),
        ]}
        for col_idx, info in COLUNAS_POSICAO.items():
            if info.get('rejeitado'):
                continue
            self.assertIn(info['tipo'], tipos_validos,
                          f'Coluna {col_idx}: tipo "{info["tipo"]}" inválido')

    def test_colunas_posicao_codigos_validos(self):
        """Todos os códigos em COLUNAS_POSICAO devem existir no COLUNAS_PRODUTO."""
        codigos_validos = set(COLUNAS_PRODUTO.values())
        for col_idx, info in COLUNAS_POSICAO.items():
            if info.get('rejeitado'):
                continue
            self.assertIn(info['codigo'], codigos_validos,
                          f'Coluna {col_idx}: código "{info["codigo"]}" não encontrado em COLUNAS_PRODUTO')

    def test_cultura_marcada_como_rejeitado(self):
        """A coluna 13 (VR Cultura) deve estar marcada como rejeitada."""
        self.assertTrue(COLUNAS_POSICAO[13].get('rejeitado'))
        self.assertEqual(COLUNAS_POSICAO[13]['codigo'], '30')
        self.assertIsNone(COLUNAS_POSICAO[13]['tipo'])

    def test_codigos_posicionais_sao_os_oficiais_da_vr(self):
        """
        Códigos conforme a tabela oficial da VR (PRODUTOS - VR.xlsx):
        cada produto tem código próprio — nada de reutilizar 207/27/28.
        """
        esperados = {
            10: '31', 11: '27', 12: '28', 13: '30', 14: '201', 15: '202',
            16: '204', 17: '243', 18: '207', 19: '209', 20: '244',
            21: '212', 22: '245', 23: '211', 24: '213', 25: '58', 26: '262',
        }
        for col, codigo in esperados.items():
            self.assertEqual(COLUNAS_POSICAO[col]['codigo'], codigo,
                             f'Coluna {col} deveria mapear para o código {codigo}')


class TestDeduplicacao(unittest.TestCase):
    """Testes para a soma de benefícios duplicados (mesmo CNPJ + CPF + produto)."""

    def test_mesmo_funcionario_mesmo_produto_soma_valores(self):
        """
        Se o mesmo CPF aparece em 2 linhas com o mesmo CNPJ e mesmo produto,
        os valores devem ser somados em uma única movimentação.
        """
        import openpyxl
        import shutil
        import tempfile

        src = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            'VR - ADM E CONDS.xlsm'
        )
        if not os.path.exists(src):
            self.skipTest(f'Arquivo de exemplo não encontrado: {src}')

        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = os.path.join(tmpdir, 'test_dedup.xlsm')
            shutil.copy(src, test_file)

            wb = openpyxl.load_workbook(test_file, data_only=False, keep_vba=True)
            ws = wb['Beneficiario']

            # Pegar um funcionário existente (linha 7) e duplicar com valor
            cpf_original = ws.cell(7, 1).value
            cnpj_original = ws.cell(7, 2).value

            # Adicionar uma nova linha com mesmo CPF e CNPJ, mesmo produto (col 10 = VR Refeição)
            nova_row = ws.max_row + 1
            ws.cell(nova_row, 1).value = cpf_original
            ws.cell(nova_row, 2).value = cnpj_original
            ws.cell(nova_row, 4).value = ws.cell(7, 4).value  # matrícula
            ws.cell(nova_row, 5).value = ws.cell(7, 5).value  # nome
            ws.cell(nova_row, 7).value = ws.cell(7, 7).value  # data nascimento

            # Colocar valor na coluna 10 (VR Refeição) tanto na linha original quanto na nova
            ws.cell(7, 10).value = 100.0
            ws.cell(nova_row, 10).value = 50.0

            wb.save(test_file)

            data = parse_fut_template(
                test_file,
                file_upload_id=1493,
                valor_max_beneficio=Decimal('9999.99'),
                administradora_cnpj='35315360000167',
            )

            # Encontrar o funcionário com o CPF duplicado
            for c in data.get('condominios', []):
                if c['cnpj'] != cnpj_original:
                    continue
                for f in c.get('funcionarios', []):
                    if f['cpf'] != cpf_original:
                        continue
                    # Deve ter apenas 1 movimentação de VR Refeição (soma de 100 + 50)
                    refeicao_movs = [
                        m for m in f['movimentacoes']
                        if m['produto'] == 'VR Refeição'
                    ]
                    self.assertEqual(len(refeicao_movs), 1,
                                     f'Esperado 1 movimentação de VR Refeição, encontrado {len(refeicao_movs)}')
                    self.assertEqual(refeicao_movs[0]['valor'], Decimal('150.00'),
                                     f'Esperado valor R$ 150.00 (100+50), encontrado {refeicao_movs[0]["valor"]}')
                    return

            self.fail(f'Funcionário CPF {cpf_original} não encontrado nos dados processados')


class TestColunasNaoMapeadas(unittest.TestCase):
    """Testes para validação de colunas sem mapeamento."""

    def test_coluna_com_valor_nao_mapeada_gera_erro(self):
        """
        Se há uma coluna fora do range 10-26 e sem header reconhecido,
        mas com valor > 0, deve gerar erro.
        """
        import openpyxl
        import shutil
        import tempfile

        src = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            'VR - ADM E CONDS.xlsm'
        )
        if not os.path.exists(src):
            self.skipTest(f'Arquivo de exemplo não encontrado: {src}')

        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = os.path.join(tmpdir, 'test_nao_mapeado.xlsm')
            shutil.copy(src, test_file)

            wb = openpyxl.load_workbook(test_file, data_only=False, keep_vba=True)
            ws = wb['Beneficiario']

            # Coluna 27 (fora do range 10-26) com valor > 0 e header não reconhecido
            ws.cell(2, 27).value = 'Produto Fantasma'
            ws.cell(7, 27).value = 999.0
            wb.save(test_file)

            data = parse_fut_template(
                test_file,
                file_upload_id=1494,
                valor_max_beneficio=Decimal('9999.99'),
                administradora_cnpj='35315360000167',
            )

            errors = data.get('errors', [])
            self.assertTrue(
                any('Coluna 27' in e and 'não está mapeada' in e for e in errors),
                f'Esperado erro de coluna não mapeada na coluna 27. Erros: {errors}'
            )


class TestDeteccaoCartaoAdmin(unittest.TestCase):
    """
    Testes da detecção de cartão admin pela aba 'Local de Entrega'.

    Além do caso clássico (1 único local), a planilha pode vir num formato
    híbrido: um local por condomínio, todos com o endereço da ADMINISTRADORA
    repetido (entrega centralizada). Esse formato também é cartão admin.
    """

    ENDERECO_ADM = ('AV ADHEMAR DE BARROS', '120', '', 'VILA SANTA ROSA', 'GUARUJÁ', 'SP', '11430003')

    def _criar_planilha(self, tmpdir, locais, beneficiarios=None):
        """
        Cria uma planilha mínima no layout VR com as abas Sumario,
        Local de Entrega e Beneficiario.
        `locais`: lista de tuplas (cnpj, nome, rua, numero, complemento,
        bairro, cidade, estado, cep).
        `beneficiarios`: lista de dicts {cpf, local, nome, nascimento, valor,
        col} (col = coluna do produto, padrão 10 = VR Refeição). Se omitido,
        cria um beneficiário válido no primeiro local.
        """
        import openpyxl

        wb = openpyxl.Workbook()
        ws_sum = wb.active
        ws_sum.title = 'Sumario'
        ws_sum.cell(1, 1).value = 'CNPJ'
        ws_sum.cell(1, 2).value = '35315360000167'

        ws_loc = wb.create_sheet('Local de Entrega')
        ws_loc.append(['Código', 'Nome', 'Tipo', 'Rua', 'Número', 'Complemento', 'Bairro', 'Cidade', 'Estado', 'CEP'])
        for cnpj, nome, rua, numero, complemento, bairro, cidade, estado, cep in locais:
            ws_loc.append([cnpj, nome, '', rua, numero, complemento, bairro, cidade, estado, cep])

        if beneficiarios is None:
            beneficiarios = [{
                'cpf': '52998224725', 'local': locais[0][0],
                'nome': 'FUNCIONARIO TESTE', 'nascimento': '01/01/1990',
                'valor': 100.0, 'col': 10,
            }]

        ws_ben = wb.create_sheet('Beneficiario')
        ws_ben.cell(2, 1).value = 'CPF'
        ws_ben.cell(2, 5).value = 'Nome'
        for i, b in enumerate(beneficiarios):
            row = 3 + i
            ws_ben.cell(row, 1).value = b['cpf']
            ws_ben.cell(row, 2).value = b.get('local', locais[0][0])
            ws_ben.cell(row, 4).value = b.get('matricula', f'MAT{row:03d}')
            ws_ben.cell(row, 5).value = b['nome']
            ws_ben.cell(row, 7).value = b.get('nascimento', '01/01/1990')
            ws_ben.cell(row, b.get('col', 10)).value = b.get('valor', 100.0)

        file_path = os.path.join(tmpdir, 'planilha_teste.xlsx')
        wb.save(file_path)
        return file_path

    def _parse(self, tmpdir, locais):
        file_path = self._criar_planilha(tmpdir, locais)
        return parse_fut_template(
            file_path,
            file_upload_id=1495,
            valor_max_beneficio=Decimal('9999.99'),
            administradora_cnpj='35315360000167',
        )

    def _local(self, cnpj, nome, endereco):
        rua, numero, complemento, bairro, cidade, estado, cep = endereco
        return (cnpj, nome, rua, numero, complemento, bairro, cidade, estado, cep)

    def test_um_unico_local_marca_cartao_admin(self):
        import tempfile
        with tempfile.TemporaryDirectory() as tmpdir:
            data = self._parse(tmpdir, [
                self._local('00034178000153', 'COND A', self.ENDERECO_ADM),
            ])
            self.assertTrue(data['cartao_admin'])

    def test_varios_locais_com_mesmo_endereco_marca_cartao_admin(self):
        """Formato híbrido: N locais, todos com o endereço da administradora."""
        import tempfile
        with tempfile.TemporaryDirectory() as tmpdir:
            data = self._parse(tmpdir, [
                self._local('00034178000153', 'COND A', self.ENDERECO_ADM),
                self._local('00034453000139', 'COND B', self.ENDERECO_ADM),
                self._local('00034680000164', 'COND C', self.ENDERECO_ADM),
            ])
            self.assertTrue(data['cartao_admin'])

    def test_mesmo_endereco_com_variacao_de_acento_e_caixa_marca_cartao_admin(self):
        import tempfile
        endereco_variante = ('av adhemar de barros', '120', '', 'Vila Santa Rosa', 'Guaruja', 'SP', '11430003')
        with tempfile.TemporaryDirectory() as tmpdir:
            data = self._parse(tmpdir, [
                self._local('00034178000153', 'COND A', self.ENDERECO_ADM),
                self._local('00034453000139', 'COND B', endereco_variante),
            ])
            self.assertTrue(data['cartao_admin'])

    def test_varios_locais_com_enderecos_distintos_nao_marca_cartao_admin(self):
        import tempfile
        endereco_b = ('RUA DAS FLORES', '55', '', 'CENTRO', 'SANTOS', 'SP', '11010000')
        with tempfile.TemporaryDirectory() as tmpdir:
            data = self._parse(tmpdir, [
                self._local('00034178000153', 'COND A', self.ENDERECO_ADM),
                self._local('00034453000139', 'COND B', endereco_b),
            ])
            self.assertFalse(data['cartao_admin'])

    def test_varios_locais_com_enderecos_vazios_nao_marca_cartao_admin(self):
        """
        Endereços todos vazios não indicam entrega centralizada — o
        faturamento consulta o CNPJ quando o endereço está vazio.
        """
        import tempfile
        endereco_vazio = ('', '', '', '', '', '', '')
        with tempfile.TemporaryDirectory() as tmpdir:
            data = self._parse(tmpdir, [
                self._local('00034178000153', 'COND A', endereco_vazio),
                self._local('00034453000139', 'COND B', endereco_vazio),
            ])
            self.assertFalse(data['cartao_admin'])


class TestCpfDuplicado(TestDeteccaoCartaoAdmin):
    """
    Mesmo CPF em mais de uma linha:
    - nome/nascimento divergentes = pessoas distintas com CPF errado →
      as DUAS linhas são bloqueadas (CPF_DUPLICADO_DIVERGENTE);
    - dados iguais = mesma pessoa → soma os valores e registra aviso
      (CPF_SOMADO) para exibição em tela.
    Reutiliza o builder de planilha da TestDeteccaoCartaoAdmin.
    """

    LOCAIS = None  # definido em setUp para reuso

    def setUp(self):
        self.locais = [self._local('00034178000153', 'COND A', self.ENDERECO_ADM)]

    def _parse_beneficiarios(self, beneficiarios):
        import tempfile
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = self._criar_planilha(tmpdir, self.locais, beneficiarios)
            return parse_fut_template(
                file_path,
                file_upload_id=1496,
                valor_max_beneficio=Decimal('9999.99'),
                administradora_cnpj='35315360000167',
            )

    def test_cpf_duplicado_divergente_bloqueia_as_duas_linhas(self):
        data = self._parse_beneficiarios([
            {'cpf': '52998224725', 'nome': 'FULANO DA SILVA', 'nascimento': '01/01/1990', 'valor': 100.0},
            {'cpf': '52998224725', 'nome': 'BELTRANO SOUZA', 'nascimento': '05/05/1985', 'valor': 200.0},
            {'cpf': '11861784775', 'nome': 'PESSOA NORMAL', 'nascimento': '02/02/1992', 'valor': 50.0},
        ])

        erros_dup = [l for l in data['linhas_com_erro'] if l['tipo_erro'] == 'CPF_DUPLICADO_DIVERGENTE']
        self.assertEqual(len(erros_dup), 2, erros_dup)

        funcionarios = [f for c in data['condominios'] for f in c['funcionarios']]
        cpfs_validos = {f['cpf'] for f in funcionarios}
        self.assertNotIn('52998224725', cpfs_validos)
        self.assertIn('11861784775', cpfs_validos)

        # Totais estornados: só a pessoa normal conta.
        self.assertEqual(data['summary']['valor_total_beneficios'], Decimal('50.0'))

    def test_cpf_duplicado_mesma_pessoa_soma_e_avisa(self):
        data = self._parse_beneficiarios([
            {'cpf': '52998224725', 'nome': 'FULANO DA SILVA', 'nascimento': '01/01/1990', 'valor': 100.0},
            {'cpf': '52998224725', 'nome': 'Fulano da Silva', 'nascimento': '01/01/1990', 'valor': 50.0},
        ])

        self.assertEqual(data['linhas_com_erro'], [])
        funcionarios = [f for c in data['condominios'] for f in c['funcionarios']]
        self.assertEqual(len(funcionarios), 1)
        self.assertEqual(funcionarios[0]['valor_bene'], Decimal('150.0'))

        avisos = data.get('avisos', [])
        self.assertEqual(len(avisos), 1, avisos)
        self.assertEqual(avisos[0]['tipo'], 'CPF_SOMADO')
        self.assertEqual(avisos[0]['cpf'], '52998224725')
        self.assertEqual(len(avisos[0]['linhas']), 2)

    def test_cpfs_distintos_seguem_normais(self):
        data = self._parse_beneficiarios([
            {'cpf': '52998224725', 'nome': 'FULANO DA SILVA', 'valor': 100.0},
            {'cpf': '11861784775', 'nome': 'PESSOA NORMAL', 'valor': 50.0},
        ])
        self.assertEqual(data['linhas_com_erro'], [])
        self.assertEqual(data.get('avisos', []), [])
        funcionarios = [f for c in data['condominios'] for f in c['funcionarios']]
        self.assertEqual(len(funcionarios), 2)


if __name__ == '__main__':
    unittest.main()
