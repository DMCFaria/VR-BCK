import io
import re
import unicodedata
from datetime import date, timedelta, datetime
from decimal import Decimal

import pandas as pd


# Mapeamento do tipo do produto para o código usado nos registros 50/60 do TXT de compra.
TIPO_PARA_CODIGO = {
    'Alimentação': '27',
    'Auto': '28',
    'Refeição': '207',
    'Multi - Home Office': '207',
    'Boas Festas': '202',
    'Multi - Alimentação': '27',
    'Multi - VR+VA': '207',
    'Multi - Refeição': '207',
    'Multi - Mobilidade': '28',
}


def _codigo_produto_por_tipo(produto):
    """Retorna o código do produto baseado no tipo, ou o código salvo se não houver mapeamento."""
    if produto.tipo:
        tipo_display = produto.get_tipo_display()
        if tipo_display in TIPO_PARA_CODIGO:
            return TIPO_PARA_CODIGO[tipo_display]
    return produto.codigo_produto


def _consultar_endereco_condominio(cnpj):
    """
    Consulta o endereço real de um condomínio pelo CNPJ.
    Usado no faturamento quando o endereço do condomínio está vazio
    (caso comum no modo cartão admin, onde a planilha traz o endereço
    da administradora como local de entrega).
    """
    from upload.services import CNPJConsultaService
    try:
        return CNPJConsultaService.consultar(cnpj, fonte="bigdatacorp_addresses")
    except Exception:
        return None


def remover_acentos(texto):
    """Remove acentos e caracteres especiais, convertendo para ASCII."""
    if not texto:
        return texto
    nfkd = unicodedata.normalize('NFKD', texto)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))
from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import views, status
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication

from entidades.models import Condominio, Funcionario, Administradora, VinculoCondominio, TaxaConfig
from beneficios.models import MovimentacaoBeneficio, Produto, Importacao, Boleto


def calcular_taxa(valor_beneficio, quantidade_dias, vinculo=None, produto=None):
    """
    Calcula a taxa de faturamento para um funcionário.

    Regras:
    - Se não houver vínculo, retorna 0
    - Busca TaxaConfig para o produto específico
    - Se não encontrar para o produto, busca configuração pelo tipo do produto
    - Se não encontrar por tipo, busca configuração genérica (produto=NULL, tipo=NULL)
    - Se não encontrar nenhuma configuração, retorna 0
    - Tipo PERC: valor_beneficio * (taxa_valor / 100)
    - Tipo FIXO: taxa_valor * quantidade_dias
    """
    if not vinculo:
        return Decimal('0.00')

    # Busca configuração para o produto específico
    taxa_config = TaxaConfig.objects.filter(
        vinculo=vinculo,
        produto=produto,
        ativo=True
    ).first()

    # Se não encontrar para o produto específico, busca pelo tipo do produto
    if not taxa_config and produto and produto.tipo:
        taxa_config = TaxaConfig.objects.filter(
            vinculo=vinculo,
            tipo=produto.tipo,
            ativo=True
        ).first()

    # Se não encontrar por tipo, busca configuração genérica
    if not taxa_config:
        taxa_config = TaxaConfig.objects.filter(
            vinculo=vinculo,
            produto__isnull=True,
            tipo__isnull=True,
            ativo=True
        ).first()

    # Se não encontrar nenhuma configuração, retorna 0
    if not taxa_config:
        return Decimal('0.00')

    # Calcula a taxa
    if taxa_config.taxa_tipo == 'PERC':
        return round(valor_beneficio * (taxa_config.taxa_valor / Decimal('100')), 2)
    else:  # FIXO
        return round(taxa_config.taxa_valor * quantidade_dias, 2)


def gerar_txt_compra(administradora_cnpj, data_competencia=None, movimentacao_ids=None):
    """
    Gera o arquivo txt_compra para envio à VR Benefícios.
    Se movimentacao_ids for passado, filtra apenas as movimentações selecionadas.
    """
    linhas = []
    seq = 1

    admin = Administradora.objects.filter(cnpj=administradora_cnpj).first()
    if not admin:
        return None, "Administradora não encontrada"

    # Header (TipoRec 00)
    linha_header = (
        f"00"  # TipoRec (2)
        f"011"  # Versao (3)
        f"{administradora_cnpj.zfill(14)}"  # CNPJ/Código Cliente (14)
        f"{remover_acentos(admin.razao_social)[:40]:<40}"  # Razão Social Cliente (40)
        f"{' ' * 282}"  # FILLER (282)
        f"{str(seq).zfill(9)}"  # Número da linha (9)
    )
    linhas.append(linha_header)
    seq += 1

    # Buscar condomínios vinculados
    query = VinculoCondominio.objects.filter(administradora=admin)
    if movimentacao_ids is not None:
        query = query.filter(
            condominio__movimentacaobeneficio__id__in=movimentacao_ids
        ).distinct()
    elif data_competencia:
        query = query.filter(
            condominio__movimentacaobeneficio__data_competencia=data_competencia
        ).distinct()

    vinculos = query.select_related('condominio').prefetch_related('gerentes')

    for vinculo in vinculos:
        condominio = vinculo.condominio

        # Buscar gerentes do vínculo
        gerentes = vinculo.gerentes.all()
        nomes_gerentes = [remover_acentos(g.nome or '') for g in gerentes if g.nome][:3]
        gerentes_str = '/'.join(nomes_gerentes)[:30] if nomes_gerentes else ''

        # Local Entrega (TipoRec 10)
        # Se cartao_admin=True, entrega na administradora; senão, no condomínio.
        endereco_entrega = admin if admin.cartao_admin else condominio
        nome_entrega = remover_acentos(endereco_entrega.razao_social if admin.cartao_admin else condominio.nome)
        numero_entrega = str(endereco_entrega.numero or '').strip()
        if not numero_entrega:
            numero_entrega = ''
        linha_local = (
            f"10"
            f"{administradora_cnpj.zfill(14)}"
            f"{condominio.cnpj[:30]:<30}"
            f"{nome_entrega[:80]:<80}"
            f"{'AVENIDA'[:20]:<20}"
            f"{remover_acentos(endereco_entrega.endereco or '')[:40]:<40}"
            f"{numero_entrega.zfill(6)[:6]:<6}"
            f"{remover_acentos(endereco_entrega.complemento or '')[:20]:<20}"
            f"{remover_acentos(endereco_entrega.bairro or '')[:30]:<30}"
            f"{remover_acentos(endereco_entrega.cidade or '')[:30]:<30}"
            f"{(endereco_entrega.estado or '')[:2]:<2}"
            f"{(endereco_entrega.cep or '').replace('-', '')[:8]:<8}"
            f"{gerentes_str[:30]:<30}"
            f"{' ' * 29}"
            f"{str(seq).zfill(9)}"
        )
        linhas.append(linha_local)
        seq += 1

        # Associação CNPJ ao Local Entrega (TipoRec 11)
        # Se cartao_admin=True, entrega na administradora; senão, no condomínio.
        nome_entrega = remover_acentos(admin.razao_social if admin.cartao_admin else condominio.nome)
        linha_assoc = (
            f"11"
            f"{administradora_cnpj.zfill(14)}"
            f"{condominio.cnpj[:30]:<30}"
            f"{administradora_cnpj.zfill(14)}"
            f"{nome_entrega[:24]:<24}"
            f"{'VR@GRUPOFEDCOPR.COM.BR'[:70]:<70}"
            f"{' ' * 187}"
            f"{str(seq).zfill(9)}"
        )
        linhas.append(linha_assoc)
        seq += 1

        # Responsáveis pelo Local de Entrega (TipoRec 12)
        emails_gerentes = [g.email.upper() for g in gerentes if g.email][:3]
        if not emails_gerentes:
            emails_gerentes = ['VR@GRUPOFEDCOPR.COM.BR']

        email1 = emails_gerentes[0][:60] if len(emails_gerentes) > 0 else ' ' * 60
        email2 = emails_gerentes[1][:60] if len(emails_gerentes) > 1 else email1
        email3 = emails_gerentes[2][:60] if len(emails_gerentes) > 2 else email1

        linha_resp = (
            f"12"
            f"{administradora_cnpj.zfill(14)}"
            f"{condominio.cnpj[:30]:<30}"
            f"{email1:<60}"
            f"{' ' * 43}"
            f"{email2:<60}"
            f"{' ' * 43}"
            f"{email3:<60}"
            f"{' ' * 29}"
            f"{str(seq).zfill(9)}"
        )
        linhas.append(linha_resp)
        seq += 1

    # Buscar movimentações
    if movimentacao_ids is not None:
        mov_query = MovimentacaoBeneficio.objects.filter(
            id__in=movimentacao_ids
        ).select_related('produto_codigo', 'funcionario_cpf', 'empresa_cnpj')
    else:
        mov_query = MovimentacaoBeneficio.objects.filter(
            empresa_cnpj__vinculocondominio__administradora=admin
        ).select_related('produto_codigo', 'funcionario_cpf', 'empresa_cnpj')

        if data_competencia:
            mov_query = mov_query.filter(data_competencia=data_competencia)

    # ⭐⭐⭐ REMOVER DUPLICAÇÃO: Apenas UM bloco de registros 30, 50 e 60 ⭐⭐⭐
    
    # Beneficiário (TipoRec 30)
    funcionarios_vistos = set()
    for mov in mov_query:
        func = mov.funcionario_cpf
        cond = mov.empresa_cnpj

        if func.cpf in funcionarios_vistos:
            continue
        funcionarios_vistos.add(func.cpf)

        data_nasc = ''
        if func.data_nascimento:
            data_nasc = func.data_nascimento.strftime('%d%m%Y')
        else:
            data_nasc = '00000000'

        linha_benef = (
            f"30"
            f"{administradora_cnpj.zfill(14)}"
            f"{func.cpf[:11]:<11}"
            f"{cond.cnpj[:30]:<30}"
            f"{' ' * 12}"
            f"CONDOMINIO"
            f"{remover_acentos(func.nome)[:40]:<40}"
            f"{' '[:24]:<24}"
            f"{data_nasc}"
            f"{' ' * 187}"
            f"{str(seq).zfill(9)}"
        )
        linhas.append(linha_benef)
        seq += 1

    # ⭐⭐⭐ APENAS UM BLOCO PARA REGISTROS 50 e 60 ⭐⭐⭐
    # Agrupamos por TIPO do produto, usando o código mapeado para o TXT.
    from itertools import groupby
    from operator import attrgetter

    def _tipo_ordenacao(mov):
        prod = mov.produto_codigo
        return prod.get_tipo_display() if prod.tipo else (prod.nome or prod.codigo_produto)

    movimentacoes_ordenadas = sorted(mov_query, key=_tipo_ordenacao)

    for tipo_display, movimentacoes_grupo in groupby(movimentacoes_ordenadas, key=_tipo_ordenacao):
        mov_list = list(movimentacoes_grupo)
        if not mov_list:
            continue

        # Pega o código do produto baseado no tipo (primeiro movimento do grupo).
        prod_cod = _codigo_produto_por_tipo(mov_list[0].produto_codigo)[:3].upper()

        # Registro 50 - Produto Voucher
        data_agend = data_competencia if data_competencia else date.today()
        linha_prod = (
            f"50"
            f"{administradora_cnpj.zfill(14)}"
            f"{prod_cod:<3}"
            f"{data_agend.strftime('%d%m%Y')}"
            f"{' ' * 314}"
            f"{str(seq).zfill(9)}"
        )
        linhas.append(linha_prod)
        seq += 1

        # Registros 60 - Benefícios Voucher
        for mov in mov_list:
            func = mov.funcionario_cpf
            valor = float(mov.valor_beneficio)
            valor_str = f"{valor:.2f}".replace('.', '').zfill(11)

            linha_beneficio = (
                f"60"
                f"{administradora_cnpj.zfill(14)}"
                f"{prod_cod:<3}"
                f"{func.cpf.zfill(11)}"
                f"{' ' * 40}"
                f"{valor_str}"
                f"{' ' * 260}"
                f"{str(seq).zfill(9)}"
            )
            linhas.append(linha_beneficio)
            seq += 1

    # Trailler (TipoRec 99)
    linha_trailler = (
        f"99"
        f"{administradora_cnpj.zfill(14)}"
        f"{' ' * 325}"
        f"{str(seq).zfill(9)}"
    )
    linhas.append(linha_trailler)

    return '\n'.join(linhas), None


def gerar_faturamento(importacao_id=None, data_inicio=None, data_fim=None, administradora_cnpj=None, condominio_cnpj=None):
    """
    Gera dados para planilha de faturamento.
    Se importacao_id for passado, filtra apenas pelas movimentações dessa importação.
    """
    query = MovimentacaoBeneficio.objects.select_related(
        'empresa_cnpj', 'funcionario_cpf', 'produto_codigo'
    ).prefetch_related(
        'empresa_cnpj__vinculocondominio_set__administradora'
    )

    if importacao_id:
        query = query.filter(importacao_id=importacao_id)
    else:
        if data_inicio:
            query = query.filter(data_competencia__gte=data_inicio)
        if data_fim:
            query = query.filter(data_competencia__lte=data_fim)
        if administradora_cnpj:
            query = query.filter(
                empresa_cnpj__vinculocondominio__administradora__cnpj=administradora_cnpj
            )
        if condominio_cnpj:
            query = query.filter(empresa_cnpj__cnpj=condominio_cnpj)

    movimentacoes = query.order_by('empresa_cnpj', 'funcionario_cpf', 'data_competencia')

    dados = []
    cache_enderecos = {}
    for mov in movimentacoes:
        func = mov.funcionario_cpf
        cond = mov.empresa_cnpj
        prod = mov.produto_codigo

        valor_unitario = mov.valor_beneficio / mov.quantidade_dias if mov.quantidade_dias > 0 else mov.valor_beneficio

        datos_periodo = mov.data_competencia.strftime('%d/%m/%Y')
        data_ini = mov.data_competencia.replace(day=1)
        data_fim = data_ini + timedelta(days=30)
        periodos = f"{data_ini.strftime('%d/%m/%Y')} - {data_fim.strftime('%d/%m/%Y')}"

        # Sempre usa o tipo do produto nos documentos exportados.
        produto_display = prod.get_tipo_display() if prod.tipo else (prod.nome or prod.codigo_produto)

        # Busca o vínculo entre o condomínio e a administradora
        vinculo = cond.vinculocondominio_set.first()
        administradora = vinculo.administradora if vinculo else None

        # Calcula a taxa
        taxa = calcular_taxa(
            valor_beneficio=mov.valor_beneficio,
            quantidade_dias=mov.quantidade_dias,
            vinculo=vinculo,
            produto=prod
        )

        # Endereço do condomínio. Se estiver vazio (comum no modo cartão admin,
        # pois a planilha traz o endereço da administradora), consulta o CNPJ.
        endereco_cond = cond.endereco or ''
        bairro_cond = cond.bairro or ''
        cidade_cond = cond.cidade or ''
        estado_cond = cond.estado or ''
        cep_cond = cond.cep or ''

        if cond.cnpj and not any([endereco_cond, bairro_cond, cidade_cond, estado_cond, cep_cond]):
            if cond.cnpj not in cache_enderecos:
                cache_enderecos[cond.cnpj] = _consultar_endereco_condominio(cond.cnpj)
            dados_cnpj = cache_enderecos[cond.cnpj]
            if dados_cnpj:
                endereco_cond = dados_cnpj.get('rua', '')
                bairro_cond = dados_cnpj.get('bairro', '')
                cidade_cond = dados_cnpj.get('cidade', '')
                estado_cond = dados_cnpj.get('estado', '')
                cep_cond = dados_cnpj.get('cep', '')

        dados.append({
            'CPF': func.cpf,
            'NOME_FUNC': func.nome,
            'PRODUTO': produto_display,
            'BENEFICIO': None,
            'CEP_FUNC': func.cep or '',
            'ENDERECO_FUNC': f'{func.endereco_rua}, {func.endereco_numero}, {func.endereco_complemento}' or '',
            'NUMERO_FUNC': func.endereco_numero or '',
            'COMPLEMENTO_FUNC': func.endereco_complemento or '',
            'BAIRRO_FUNC': func.endereco_bairro or '',
            'VALOR_UNITARIO': float(valor_unitario),
            'QUANTIDADE': mov.quantidade_dias,
            'VALOR_RECARGA_BENE': float(mov.valor_beneficio),
            'REPASSE_VT': None,
            'DEPARTAMENTO': cond.nome,
            'CNPJ': cond.cnpj,
            'ENDERECO': endereco_cond,
            'BAIRRO': bairro_cond,
            'CIDADE': cidade_cond,
            'UF': estado_cond,
            'CEP': cep_cond,
            'TAXA': float(taxa),
            'vencimento': datos_periodo,
            'periodos': periodos.split('-')[0],
            'periodo2': periodos.split('-')[1]
        })
    return dados


class GetImportacaoSelectDataView(views.APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request, importacao_id):
        try:
            importacao = Importacao.objects.get(id=importacao_id)
        except Importacao.DoesNotExist:
            return Response({'detail': 'Importação não encontrada.'}, status=404)

        boletos = Boleto.objects.filter(faturamento_id=importacao_id)

        fedhub_status_map = {}
        faturas_num = list(set(b.fatura for b in boletos if b.fatura))
        if faturas_num:
            try:
                from core.fedhub.services.fedhub_service import FedhubService
                fedhub = FedhubService()
                for fat_num in faturas_num:
                    fedhub_boletos = fedhub.buscar_todos_boletos_por_fatura(fat_num)
                    if isinstance(fedhub_boletos, list):
                        for fb in fedhub_boletos:
                            doc = fb.get("documento")
                            if doc:
                                fedhub_status_map[str(doc).strip()] = {
                                    "status": fb.get("status"),
                                    "baixa": bool(fb.get("baixa", False)),
                                    "dt_baixa": fb.get("dt_baixa"),
                                }
            except Exception as e:
                import logging
                logging.warning(f"Erro ao buscar status no FedHub para select-data: {e}")

        boletos_data = []
        for b in boletos:
            doc_key = str(b.documento).strip() if b.documento else None
            fh = fedhub_status_map.get(doc_key, {}) if doc_key else {}

            boletos_data.append({
                "id": b.id,
                "documento": b.documento,
                "cnpj_cobrado": b.cnpj_cobrado,
                "nome_cobrado": b.nome_cobrado,
                "valor": float(b.valor) if b.valor else 0.0,
                "vencimento": b.vencimento.strftime('%Y-%m-%d') if b.vencimento else None,
                "baixa": fh.get("baixa", b.baixa),
                "dt_baixa": fh.get("dt_baixa", b.dt_baixa.strftime('%Y-%m-%d') if b.dt_baixa else None),
                "status": fh.get("status", b.status),
            })

        movs = MovimentacaoBeneficio.objects.filter(importacao=importacao).select_related(
            'empresa_cnpj', 'funcionario_cpf', 'produto_codigo'
        ).order_by('empresa_cnpj__nome', 'funcionario_cpf__nome')

        condominios_dict = {}
        for m in movs:
            cond_cnpj = m.empresa_cnpj.cnpj
            if cond_cnpj not in condominios_dict:
                condominios_dict[cond_cnpj] = {
                    "cnpj": cond_cnpj,
                    "nome": m.empresa_cnpj.nome,
                    "funcionarios": {}
                }

            func_cpf = m.funcionario_cpf.cpf
            if func_cpf not in condominios_dict[cond_cnpj]["funcionarios"]:
                condominios_dict[cond_cnpj]["funcionarios"][func_cpf] = {
                    "cpf": func_cpf,
                    "nome": m.funcionario_cpf.nome,
                    "movimentacoes": []
                }

            condominios_dict[cond_cnpj]["funcionarios"][func_cpf]["movimentacoes"].append({
                "id": m.id,
                "produto_codigo": m.produto_codigo.codigo_produto,
                "produto_nome": m.produto_codigo.get_tipo_display() if m.produto_codigo.tipo else m.produto_codigo.nome,
                "valor_beneficio": float(m.valor_beneficio),
                "quantidade_dias": m.quantidade_dias,
                "data_competencia": m.data_competencia.strftime('%Y-%m-%d') if m.data_competencia else None,
            })

        condominios_list = []
        for cond_cnpj, cond_data in condominios_dict.items():
            funcs_list = list(cond_data["funcionarios"].values())
            cond_data["funcionarios"] = funcs_list
            condominios_list.append(cond_data)

        return Response({
            "importacao_id": importacao.id,
            "condominios": condominios_list,
            "boletos": boletos_data,
        })


class ListarTodosBoletosView(views.APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        from collections import defaultdict
        import logging
        log = logging.getLogger(__name__)

        boletos = Boleto.objects.select_related('faturamento', 'faturamento__importacao', 'faturamento__administradora').all()

        grupos = defaultdict(list)
        for b in boletos:
            if b.fatura:
                grupos[b.fatura].append(b)

        faturas_ordenadas = sorted(grupos.keys(), reverse=True)

        fedhub_status_map = {}
        try:
            from core.fedhub.services.fedhub_service import FedhubService
            fedhub = FedhubService()
            for fat_num in faturas_ordenadas:
                fedhub_boletos = fedhub.buscar_todos_boletos_por_fatura(fat_num)
                if isinstance(fedhub_boletos, list):
                    for fb in fedhub_boletos:
                        doc = fb.get("documento")
                        if doc:
                            fedhub_status_map[str(doc).strip()] = {
                                "status": fb.get("status"),
                                "baixa": bool(fb.get("baixa", False)),
                                "dt_baixa": fb.get("dt_baixa"),
                            }
        except Exception as e:
            log.warning(f"Erro ao buscar status no FedHub: {e}")

        boletos_data = []
        for fat_num in faturas_ordenadas:
            for b in grupos[fat_num]:
                doc_key = str(b.documento).strip() if b.documento else None
                fh = fedhub_status_map.get(doc_key, {}) if doc_key else {}
                boletos_data.append({
                    "fatura": b.fatura,
                    "documento": b.documento,
                    "cnpj_cobrado": b.cnpj_cobrado,
                    "nome_cobrado": b.nome_cobrado,
                    "valor": float(b.valor) if b.valor else 0.0,
                    "vencimento": b.vencimento.strftime('%Y-%m-%d') if b.vencimento else None,
                    "baixa": fh.get("baixa", b.baixa),
                    "dt_baixa": fh.get("dt_baixa", b.dt_baixa.strftime('%Y-%m-%d') if b.dt_baixa else None),
                    "status": fh.get("status", b.status),
                    "nosso_numero": b.nosso_numero,
                })

        return Response(boletos_data)


class ExportTxtCompraView(views.APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def post(self, request, *args, **kwargs):
        importacao_id = request.data.get('importacao_id')
        data_competencia_str = request.data.get('data_competencia')
        movimentacao_ids = request.data.get('movimentacao_ids')

        if not importacao_id:
            return Response(
                {'detail': 'Parâmetro importacao_id é obrigatório.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            importacao = Importacao.objects.get(id=importacao_id)
        except Importacao.DoesNotExist:
            return Response(
                {'detail': 'Importação não encontrada.'},
                status=status.HTTP_404_NOT_FOUND
            )

        if not importacao.administradora:
            return Response(
                {'detail': 'Importação não possui administradora vinculada.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        administradora_cnpj = importacao.administradora.cnpj

        data_competencia = None
        if data_competencia_str:
            try:
                data_competencia = datetime.strptime(data_competencia_str, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {'detail': 'Formato de data_competencia inválido. Use YYYY-MM-DD.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            first_mov = importacao.movimentacoes.first()
            if first_mov:
                data_competencia = first_mov.data_competencia
            else:
                data_competencia = date.today()

        txt_content, error = gerar_txt_compra(administradora_cnpj, data_competencia, movimentacao_ids)

        if error:
            return Response(
                {'detail': error},
                status=status.HTTP_400_BAD_REQUEST
            )

        response = HttpResponse(txt_content.encode('latin-1', errors='replace'), content_type='text/plain; charset=iso-8859-1')
        response['Content-Disposition'] = f'attachment; filename="PEDIDO_VR_{date.today().strftime("%Y%m%d")}.txt"'
        return response


def gerar_planilha_compra_vt(importacao_id):
    from beneficios.models import Importacao, MovimentacaoBeneficio, Produto
    from entidades.models import Condominio, Funcionario
    from django.db.models import Sum
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    try:
        importacao = Importacao.objects.select_related('administradora', 'file_upload').get(id=importacao_id)
    except Importacao.DoesNotExist:
        return None, "Importação não encontrada"

    movimentacoes = MovimentacaoBeneficio.objects.filter(
        importacao_id=importacao_id
    ).select_related(
        'funcionario_cpf', 'empresa_cnpj', 'produto_codigo'
    ).order_by('empresa_cnpj', 'funcionario_cpf')

    if not movimentacoes:
        return None, "Nenhuma movimentação encontrada"

    admin_cnpj = importacao.administradora.cnpj if importacao.administradora else ''

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'USUARIOS'

    # Styles
    font_header = Font(name='Arial', size=10, bold=True)
    font_section = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    fill_section = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    font_data = Font(name='Arial', size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Fixed columns (old template layout, no department address)
    FIXED_COLS = 23
    ITEM_BLOCKS = 10
    ITEM_COLS_PER_BLOCK = 4
    TOTAL_COLS = FIXED_COLS + (ITEM_BLOCKS * ITEM_COLS_PER_BLOCK)

    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    cell = ws.cell(row=1, column=1, value='CADASTRO DE USUÁRIOS')
    cell.font = Font(name='Arial', size=14, bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Row 2: Required fields note
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
    cell = ws.cell(row=2, column=1, value='*CAMPOS OBRIGATÓRIOS')
    cell.font = Font(name='Arial', size=9, italic=True, color='FF0000')

    # Row 3: Section headers
    section_headers = {
        1: 'IDENTIFICAÇÃO DO USUÁRIO',
        9: 'DADOS DO USUÁRIO',
        17: 'ENDEREÇO DO USUÁRIO',
    }
    merge_ranges = {
        1: (1, 8),
        9: (9, 16),
        17: (17, 23),
    }
    for col, text in section_headers.items():
        cell = ws.cell(row=3, column=col, value=text)
        cell.font = font_section
        cell.fill = fill_section
        cell.alignment = Alignment(horizontal='center')
        start_c, end_c = merge_ranges[col]
        ws.merge_cells(start_row=3, start_column=start_c, end_row=3, end_column=end_c)

    item_start = FIXED_COLS + 1
    for item_num in range(ITEM_BLOCKS):
        col = item_start + (item_num * ITEM_COLS_PER_BLOCK)
        cell = ws.cell(row=3, column=col, value=f'ITEM {item_num + 1}')
        cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(
            start_row=3, start_column=col,
            end_row=3, end_column=col + 3
        )

    # Row 4: Column headers
    col_headers = [
        ('CNPJ*', 1), ('MATRÍCULA*', 2), ('NOME COMPLETO*', 3),
        ('EMAIL', 4), ('CELULAR', 5), ('ATIVO', 6),
        ('ENDEREÇO*', 7), ('CARGO', 8),
        ('DEPARTAMENTO', 9),
        ('DIAS TRABALHADOS*', 10), ('CPF*', 11),
        ('RG.', 12), ('DG.', 13), ('EST.RG', 14),
        ('DATA DE NASCIMENTO', 15), ('NOME DA MÃE', 16),
        ('LOGRADOURO', 17), ('NÚMERO', 18), ('COMPLEMENTO', 19),
        ('BAIRRO', 20), ('CEP', 21), ('CIDADE', 22), ('ESTADO', 23),
    ]
    for text, col in col_headers:
        cell = ws.cell(row=4, column=col, value=text)
        cell.font = font_header
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    item_headers = ['CÓD.', 'QTD.', 'DIAS.', 'VALOR']
    for item_num in range(ITEM_BLOCKS):
        base_col = item_start + (item_num * ITEM_COLS_PER_BLOCK)
        for i, h in enumerate(item_headers):
            cell = ws.cell(row=4, column=base_col + i, value=h)
            cell.font = font_header
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    # Group movimentacoes by employee (CPF + CNPJ)
    from collections import OrderedDict
    emp_map = OrderedDict()

    for mov in movimentacoes:
        func = mov.funcionario_cpf
        cond = mov.empresa_cnpj
        prod = mov.produto_codigo
        key = f"{cond.cnpj}_{func.cpf}"

        if key not in emp_map:
            emp_map[key] = {
                'cnpj_cond': cond.cnpj,
                'nome_cond': cond.nome,
                'matricula': func.matricula or '',
                'nome': func.nome or '',
                'cpf': func.cpf,
                'funcao': func.funcao or '',
                'data_nascimento': func.data_nascimento,
                'logradouro': func.endereco_rua or '',
                'numero': func.endereco_numero or '',
                'complemento': func.endereco_complemento or '',
                'bairro': func.endereco_bairro or '',
                'cep': func.cep or '',
                'dias_trabalhados': 0,
                'itens': {},
            }

        dias = mov.quantidade_dias if mov.quantidade_dias else 0
        emp_map[key]['dias_trabalhados'] = max(emp_map[key]['dias_trabalhados'], dias)

        cod_prod = (prod.codigo_produto or '').strip()[:50]
        qtd = max(1, dias)  # quantidade = dias for VT
        if cod_prod not in emp_map[key]['itens']:
            emp_map[key]['itens'][cod_prod] = {
                'codigo': cod_prod,
                'quantidade': 0,
                'dias': 0,
                'valor_total': 0,
            }
        emp_map[key]['itens'][cod_prod]['quantidade'] += 1
        emp_map[key]['itens'][cod_prod]['dias'] += dias
        emp_map[key]['itens'][cod_prod]['valor_total'] += float(mov.valor_beneficio)

    row_num = 5
    admin_cnpj_clean = re.sub(r'[^0-9]', '', admin_cnpj).zfill(14)[:14]

    for emp_key, emp_data in emp_map.items():
        itens_list = list(emp_data['itens'].values())

        data_nasc = ''
        if emp_data['data_nascimento']:
            if isinstance(emp_data['data_nascimento'], date):
                data_nasc = emp_data['data_nascimento'].strftime('%d/%m/%Y')
            else:
                data_nasc = str(emp_data['data_nascimento'])

        data_row = [
            admin_cnpj_clean,                                          # 1: CNPJ*
            emp_data['matricula'],                                     # 2: MATRÍCULA*
            emp_data['nome'],                                          # 3: NOME COMPLETO*
            '',                                                         # 4: EMAIL
            '',                                                         # 5: CELULAR
            'SIM',                                                      # 6: ATIVO
            emp_data['logradouro'],                                    # 7: ENDEREÇO*
            emp_data['funcao'],                                        # 8: CARGO
            f"{emp_data['cnpj_cond']} - {emp_data['nome_cond']}",      # 9: DEPARTAMENTO
            emp_data['dias_trabalhados'],                               # 10: DIAS TRABALHADOS*
            emp_data['cpf'],                                           # 11: CPF*
            '',                                                         # 12: RG.
            '',                                                         # 13: DG.
            '',                                                         # 14: EST.RG
            data_nasc,                                                  # 15: DATA DE NASCIMENTO
            '',                                                         # 16: NOME DA MÃE
            emp_data['logradouro'],                                     # 17: LOGRADOURO
            emp_data['numero'],                                         # 18: NÚMERO
            emp_data['complemento'],                                    # 19: COMPLEMENTO
            emp_data['bairro'],                                         # 20: BAIRRO
            emp_data['cep'],                                            # 21: CEP
            '',                                                         # 22: CIDADE
            '',                                                         # 23: ESTADO
        ]

        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = font_data
            cell.border = thin_border

        # Items
        for item_idx in range(ITEM_BLOCKS):
            base_col = item_start + (item_idx * ITEM_COLS_PER_BLOCK)
            if item_idx < len(itens_list):
                item = itens_list[item_idx]
                ws.cell(row=row_num, column=base_col, value=item['codigo']).font = font_data
                ws.cell(row=row_num, column=base_col + 1, value=item['quantidade']).font = font_data
                ws.cell(row=row_num, column=base_col + 2, value=item['dias']).font = font_data
                valor_cell = ws.cell(row=row_num, column=base_col + 3, value=round(item['valor_total'], 2))
                valor_cell.font = font_data
                valor_cell.number_format = '#,##0.00'
            else:
                for i in range(4):
                    ws.cell(row=row_num, column=base_col + i, value='').font = font_data

            for i in range(4):
                ws.cell(row=row_num, column=base_col + i).border = thin_border

        row_num += 1

    # Column widths
    col_widths = {1: 18, 2: 15, 3: 35, 4: 30, 5: 15, 6: 8, 7: 40, 8: 20,
                  9: 45, 10: 15, 11: 15, 12: 10, 13: 8, 14: 8,
                  15: 18, 16: 30, 17: 35, 18: 10, 19: 20, 20: 25,
                  21: 15, 22: 25, 23: 10}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    item_width = 10
    for item_num in range(ITEM_BLOCKS):
        base_col = item_start + (item_num * ITEM_COLS_PER_BLOCK)
        ws.column_dimensions[get_column_letter(base_col)].width = item_width
        ws.column_dimensions[get_column_letter(base_col + 1)].width = item_width
        ws.column_dimensions[get_column_letter(base_col + 2)].width = item_width
        ws.column_dimensions[get_column_letter(base_col + 3)].width = item_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, None


class ExportVTCompraView(views.APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request, *args, **kwargs):
        importacao_id = request.query_params.get('importacao_id')

        if not importacao_id:
            return Response(
                {'detail': 'Parâmetro importacao_id é obrigatório.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from beneficios.models import Importacao
        try:
            importacao = Importacao.objects.get(id=importacao_id)
        except Importacao.DoesNotExist:
            return Response(
                {'detail': 'Importação não encontrada.'},
                status=status.HTTP_404_NOT_FOUND
            )

        if importacao.modelo_importacao != 'VT-AUTO':
            return Response(
                {'detail': 'Esta importação não é do tipo VT.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        output, error = gerar_planilha_compra_vt(importacao_id)

        if error:
            return Response({'detail': error}, status=status.HTTP_400_BAD_REQUEST)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"PEDIDO_VT_{importacao_id}_{date.today().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ExportFaturamentoView(views.APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request, *args, **kwargs):
        importacao_id = request.query_params.get('importacao_id')

        if not importacao_id:
            return Response(
                {'detail': 'Parâmetro importacao_id é obrigatório.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from beneficios.models import Importacao

        try:
            importacao = Importacao.objects.get(id=importacao_id)
        except Importacao.DoesNotExist:
            return Response(
                {'detail': 'Importação não encontrada.'},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            dados = gerar_faturamento(importacao_id=importacao_id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erro ao gerar dados de faturamento: {str(e)}", exc_info=True)
            return Response(
                {'detail': f'Erro ao gerar dados de faturamento: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        if not dados:
            return Response(
                {'detail': 'Nenhuma movimentação encontrada para esta importação.'},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            df = pd.DataFrame(dados)
            
            # Verificar se o DataFrame não está vazio
            if df.empty:
                return Response(
                    {'detail': 'Nenhum dado para exportar.'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Definir colunas
            columns_order = [
                'CPF', 'NOME_FUNC', 'PRODUTO', 'BENEFICIO', 'VALOR_UNITARIO',
                'QUANTIDADE', 'VALOR_RECARGA_BENE', 'REPASSE_VT', 'DEPARTAMENTO',
                'CNPJ', 'ENDERECO', 'BAIRRO', 'CIDADE', 'UF', 'CEP',
                'TAXA', 'vencimento', 'periodos', 'periodo2'
            ]
            
            # Filtrar apenas colunas que existem
            existing_columns = [c for c in columns_order if c in df.columns]
            df = df[existing_columns]

            output = io.BytesIO()
            
            # Tentar com xlsxwriter primeiro (mais estável)
            try:
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    df.to_excel(writer, index=False, sheet_name='Faturamento')
            except Exception as e:
                # Fallback para openpyxl
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Faturamento')
            
            output.seek(0)
            
            # Verificar se o buffer não está vazio
            if output.getbuffer().nbytes == 0:
                return Response(
                    {'detail': 'Erro ao gerar arquivo Excel: buffer vazio.'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Usar filename com .xlsx e aspas para evitar problemas
            filename = f"PLAN_FATURAMENTO_{importacao_id}_{date.today().strftime('%Y%m%d')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['Pragma'] = 'no-cache'
            response['Expires'] = '0'
            
            return response
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erro ao gerar Excel: {str(e)}", exc_info=True)
            return Response(
                {'detail': f'Erro ao gerar planilha: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )