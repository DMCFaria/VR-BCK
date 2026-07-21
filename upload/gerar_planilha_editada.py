import io
import os
import logging
from datetime import datetime
from decimal import Decimal

import openpyxl

logger = logging.getLogger(__name__)

COLUNAS_PRODUTO = {
    'Refeição': '207', 'Multi Refeição': '207', 'Alimentação': '27',
    'Multi Alimentação': '27', 'Auto': '28', 'Mobilidade': '28',
    'VR Mobilidade': '28', 'Multi Mobilidade': '28', 'VR Multi Mobilidade': '28',
    'Cesta': '201', 'Boas Festas': '202', 'Auxílio Alimentação': '204',
    'Multi Auxílio Alimentação': '204', 'Auxílio Refeição': '207',
    'Multi Auxílio Refeição': '207', 'Multibenefício': '207', 'Multibenefícios': '207',
    'Auxílio VR+VA': '207', 'Multi Auxílio VR+VA': '207', 'Multi Premiação': '207',
    'VR Refeição': '207', 'VR Alimentação': '27', 'VR Auto': '28',
    'VR Alimentação Cesta': '201', 'VR Boas Festas': '202', 'VR Auxílio Alimentação': '204',
    'VR Auxílio Refeição': '207', 'VR Multibenefícios': '207', 'VR+VA': '207',
    'VR Multi Refeição': '207', 'VR Multi Alimentação': '27',
    'VR Multi Alimentação Valor do crédito': '27',
    'VR Multi Refeição Auxílio': '207', 'VR Multi Alimentação Auxílio': '204',
    'VR Multi VR+VA': '207',
}


def editar_planilha_original(file_path, dados_modificados, data_competencia=None):
    """
    Edita a planilha original preservando formatação e macros.

    Abre o arquivo original, limpa as linhas de dados e preenche
    com os dados modificados, mantendo headers, formatação e macros VBA.

    Retorna: BytesIO com o arquivo editado
    """
    logger.info(f"[EDITAR_PLANILHA] Iniciando edição da planilha original: {file_path}")
    logger.info(f"[EDITAR_PLANILHA] dados_modificados presente: {bool(dados_modificados)}")
    logger.debug(f"[EDITAR_PLANILHA] data_competencia: {data_competencia}")

    if not dados_modificados:
        logger.warning("[EDITAR_PLANILHA] dados_modificados está vazio ou nulo")
        return None

    if not os.path.exists(file_path):
        logger.error(f"[EDITAR_PLANILHA] Arquivo não encontrado: {file_path}")
        return None

    try:
        condominios = dados_modificados.get('condominios', [])
        total_funcionarios = sum(len(c.get('funcionarios', [])) for c in condominios)
        logger.info(f"[EDITAR_PLANILHA] Dados recebidos - condominios: {len(condominios)}, funcionarios: {total_funcionarios}")

        file_ext = os.path.splitext(file_path)[1].lower()
        keep_vba = file_ext == '.xlsm'
        logger.debug(f"[EDITAR_PLANILHA] Extensão do arquivo: {file_ext}, keep_vba: {keep_vba}")

        wb = _limpar_dimensoes_extras(file_path, keep_vba=keep_vba)
        logger.info(f"[EDITAR_PLANILHA] Planilha aberta e limpa - sheets: {wb.sheetnames}")

        _editar_sheet_sumario(wb, dados_modificados, data_competencia)
        _editar_sheet_local_entrega(wb, dados_modificados)
        _editar_sheet_beneficiario(wb, dados_modificados)

        output = io.BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)

        logger.info(f"[EDITAR_PLANILHA] Planilha editada com sucesso - tamanho: {output.getbuffer().nbytes} bytes")
        return output

    except Exception as e:
        logger.error(f"[EDITAR_PLANILHA] Erro ao editar planilha: {str(e)}", exc_info=True)
        return None


def editar_planilha_vt(file_path, dados_modificados, data_competencia=None):
    """
    Edita a planilha de Vale Transporte (VT) com os dados validados.

    Atualiza as abas EMPRESA e USUARIOS preservando o layout do template VT.

    Retorna: BytesIO com o arquivo editado
    """
    logger.info(f"[EDITAR_PLANILHA_VT] Iniciando edição da planilha VT: {file_path}")
    logger.info(f"[EDITAR_PLANILHA_VT] dados_modificados presente: {bool(dados_modificados)}")

    if not dados_modificados:
        logger.warning("[EDITAR_PLANILHA_VT] dados_modificados está vazio ou nulo")
        return None

    if not os.path.exists(file_path):
        logger.error(f"[EDITAR_PLANILHA_VT] Arquivo não encontrado: {file_path}")
        return None

    try:
        condominios = dados_modificados.get('condominios', [])
        total_funcionarios = sum(len(c.get('funcionarios', [])) for c in condominios)
        logger.info(f"[EDITAR_PLANILHA_VT] Dados recebidos - condominios: {len(condominios)}, funcionarios: {total_funcionarios}")

        file_ext = os.path.splitext(file_path)[1].lower()
        keep_vba = file_ext == '.xlsm'
        logger.debug(f"[EDITAR_PLANILHA_VT] Extensão do arquivo: {file_ext}, keep_vba: {keep_vba}")

        wb = _limpar_dimensoes_extras(file_path, keep_vba=keep_vba)
        logger.info(f"[EDITAR_PLANILHA_VT] Planilha aberta e limpa - sheets: {wb.sheetnames}")

        _editar_sheet_empresa_vt(wb, condominios)
        _editar_sheet_usuarios_vt(wb, condominios)

        output = io.BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)

        logger.info(f"[EDITAR_PLANILHA_VT] Planilha VT editada com sucesso - tamanho: {output.getbuffer().nbytes} bytes")
        return output

    except Exception as e:
        logger.error(f"[EDITAR_PLANILHA_VT] Erro ao editar planilha VT: {str(e)}", exc_info=True)
        return None


def _editar_sheet_empresa_vt(wb, condominios):
    if 'EMPRESA' not in wb.sheetnames:
        logger.warning("[EDITAR_PLANILHA_VT] Sheet 'EMPRESA' não encontrada")
        return

    ws = wb['EMPRESA']
    logger.debug("[EDITAR_PLANILHA_VT] Editando sheet 'EMPRESA'")

    if not condominios:
        logger.warning("[EDITAR_PLANILHA_VT] Nenhum condomínio para atualizar EMPRESA")
        return

    condo = condominios[0]
    cnpj = ''.join(filter(str.isdigit, str(condo.get('cnpj', ''))))
    nome = condo.get('nome', '')
    logradouro = condo.get('rua', '')
    numero = condo.get('numero', '')
    complemento = condo.get('complemento', '')
    cep = ''.join(filter(str.isdigit, str(condo.get('cep', ''))))
    bairro = condo.get('bairro', '')
    cidade = condo.get('cidade', '')
    estado = condo.get('estado', '')

    # Linha 5 do template: CNPJ, EMPRESA, CÓDIGO, LOGRADOURO, NÚMERO, COMPLEMENTO, CEP, BAIRRO, CIDADE, ESTADO, NOME
    _set_cell_value(ws, row=5, column=1, value=cnpj)
    _set_cell_value(ws, row=5, column=2, value=nome)
    _set_cell_value(ws, row=5, column=4, value=logradouro)
    _set_cell_value(ws, row=5, column=5, value=numero)
    _set_cell_value(ws, row=5, column=6, value=complemento)
    _set_cell_value(ws, row=5, column=7, value=cep)
    _set_cell_value(ws, row=5, column=8, value=bairro)
    _set_cell_value(ws, row=5, column=9, value=cidade)
    _set_cell_value(ws, row=5, column=10, value=estado)
    _set_cell_value(ws, row=5, column=11, value=f"Local de Entrega {condo.get('codigo', '')}".strip())

    logger.info(f"[EDITAR_PLANILHA_VT] EMPRESA atualizada: cnpj={cnpj}, nome={nome}")


def _editar_sheet_usuarios_vt(wb, condominios):
    if 'USUARIOS' not in wb.sheetnames:
        logger.warning("[EDITAR_PLANILHA_VT] Sheet 'USUARIOS' não encontrada")
        return

    ws = wb['USUARIOS']
    logger.debug("[EDITAR_PLANILHA_VT] Editando sheet 'USUARIOS'")

    # Encontra a linha do cabeçalho (CNPJ*)
    header_row_idx = None
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val and str(val).strip().upper() in ('CNPJ*', 'CNPJ'):
            header_row_idx = row_idx
            break

    if header_row_idx is None:
        logger.warning("[EDITAR_PLANILHA_VT] Cabeçalho 'CNPJ*' não encontrado em USUARIOS")
        return

    logger.debug(f"[EDITAR_PLANILHA_VT] Cabeçalho USUARIOS encontrado na linha {header_row_idx}")

    # Remove todas as linhas de dados abaixo do cabeçalho
    max_row = ws.max_row
    if max_row > header_row_idx:
        ws.delete_rows(header_row_idx + 1, max_row - header_row_idx)
        logger.debug(f"[EDITAR_PLANILHA_VT] USUARIOS - {max_row - header_row_idx} linhas antigas removidas")

    # Mapeamento de colunas conforme Template_VT.xlsx
    row_num = header_row_idx + 1
    for condo in condominios:
        cnpj_empresa = ''.join(filter(str.isdigit, str(condo.get('cnpj', ''))))
        cnpj_depto = cnpj_empresa
        nome_depto = condo.get('nome', '')
        departamento = f"{cnpj_depto} - {nome_depto}" if cnpj_depto and nome_depto else nome_depto

        for func in condo.get('funcionarios', []):
            cpf = ''.join(filter(str.isdigit, str(func.get('cpf', ''))))
            matricula = func.get('matricula', '')
            nome = func.get('nome', '')
            cargo = func.get('funcao', '')
            dias_trabalhados = 0

            # Endereço do funcionário
            logradouro = func.get('endereco_rua', '') or func.get('logradouro', '')
            numero = func.get('endereco_numero', '') or func.get('numero', '')
            complemento = func.get('endereco_complemento', '') or func.get('complemento', '')
            bairro = func.get('endereco_bairro', '') or func.get('bairro', '')
            cep = ''.join(filter(str.isdigit, str(func.get('cep', '') or condo.get('cep', ''))))
            cidade = func.get('cidade', '') or condo.get('cidade', '')
            estado = func.get('estado', '') or condo.get('estado', '')

            # Endereço do departamento (condomínio)
            endereco_depto = condo.get('rua', '')
            numero_depto = condo.get('numero', '')
            complemento_depto = condo.get('complemento', '')
            bairro_depto = condo.get('bairro', '')
            cep_depto = ''.join(filter(str.isdigit, str(condo.get('cep', ''))))
            cidade_depto = condo.get('cidade', '')
            estado_depto = condo.get('estado', '')

            data_nasc = func.get('data_nascimento', '')
            if data_nasc and isinstance(data_nasc, str):
                try:
                    dt = datetime.strptime(data_nasc, '%Y-%m-%d')
                    data_nasc = dt.strftime('%d/%m/%Y')
                except ValueError:
                    pass
            elif data_nasc and hasattr(data_nasc, 'strftime'):
                data_nasc = data_nasc.strftime('%d/%m/%Y')

            _set_cell_value(ws, row=row_num, column=1, value=cnpj_empresa)
            _set_cell_value(ws, row=row_num, column=2, value=matricula)
            _set_cell_value(ws, row=row_num, column=3, value=nome)
            _set_cell_value(ws, row=row_num, column=6, value='ATIVO')
            _set_cell_value(ws, row=row_num, column=7, value=f"{condo.get('codigo', '')}")
            _set_cell_value(ws, row=row_num, column=8, value=cargo)
            _set_cell_value(ws, row=row_num, column=9, value=departamento)
            _set_cell_value(ws, row=row_num, column=10, value=cep_depto)
            _set_cell_value(ws, row=row_num, column=11, value=cidade_depto)
            _set_cell_value(ws, row=row_num, column=12, value=bairro_depto)
            _set_cell_value(ws, row=row_num, column=13, value=estado_depto)
            _set_cell_value(ws, row=row_num, column=14, value=endereco_depto)
            _set_cell_value(ws, row=row_num, column=15, value=dias_trabalhados)
            _set_cell_value(ws, row=row_num, column=16, value=cpf)
            _set_cell_value(ws, row=row_num, column=20, value=data_nasc)
            _set_cell_value(ws, row=row_num, column=22, value=logradouro)
            _set_cell_value(ws, row=row_num, column=23, value=numero)
            _set_cell_value(ws, row=row_num, column=24, value=complemento)
            _set_cell_value(ws, row=row_num, column=25, value=bairro)
            _set_cell_value(ws, row=row_num, column=26, value=cep)
            _set_cell_value(ws, row=row_num, column=27, value=cidade)
            _set_cell_value(ws, row=row_num, column=28, value=estado)

            logger.debug(f"[EDITAR_PLANILHA_VT] USUARIOS - linha {row_num}: cnpj={cnpj_empresa}, matricula={matricula}, nome={nome}, cpf={cpf}")

            # Itens de VT (movimentacoes)
            # Como o VALOR já é o total do item, mantemos DIAS=1 para que o
            # parser VT calcule corretamente: QTD x 1 x VALOR = valor total.
            movimentacoes = func.get('movimentacoes', [])
            for item_idx, mov in enumerate(movimentacoes[:10], start=1):
                base_col = 29 + ((item_idx - 1) * 4)
                codigo = mov.get('codigo_produto', '')
                quantidade = mov.get('quantidade', 1) or 1
                dias = 1
                valor = mov.get('valor', 0) or mov.get('valor_beneficio_total', 0)

                _set_cell_value(ws, row=row_num, column=base_col, value=codigo)
                _set_cell_value(ws, row=row_num, column=base_col + 1, value=quantidade)
                _set_cell_value(ws, row=row_num, column=base_col + 2, value=dias)
                _set_cell_value(ws, row=row_num, column=base_col + 3, value=valor)
                logger.debug(f"[EDITAR_PLANILHA_VT] USUARIOS - item {item_idx} na coluna {base_col}: codigo={codigo}, qtd={quantidade}, dias={dias}, valor={valor}")

            row_num += 1

    logger.info(f"[EDITAR_PLANILHA_VT] USUARIOS atualizado: {row_num - header_row_idx - 1} funcionarios")


def _limpar_dimensoes_extras(file_path, keep_vba=False):
    """
    Remove linhas e colunas vazias além da área real de dados em cada worksheet.

    Planilhas com formatação/estilo residual em colunas/linhas distantes
    (ex: max_col=16378 com dados apenas até a coluna 17) consomem muita RAM
    e aumentam o tempo de processamento. Esta função usa read_only para
    descobrir as dimensões reais e depois abre em modo normal apenas para
    deletar o excesso, economizando memória.

    Retorna o workbook aberto em modo normal e já limpo.
    """
    # 1º passo: descobrir dimensões reais com baixo consumo de memória (read_only)
    dimensoes = {}
    wb_read = openpyxl.load_workbook(file_path, read_only=True, keep_vba=keep_vba)
    try:
        for sheet_name in wb_read.sheetnames:
            ws = wb_read[sheet_name]

            last_row = 0
            last_col = 0

            # Itera uma única vez sobre todas as células para encontrar
            # a última linha e coluna com dados reais.
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                for col_idx, value in enumerate(row, start=1):
                    if value is not None and str(value).strip() != '':
                        if row_idx > last_row:
                            last_row = row_idx
                        if col_idx > last_col:
                            last_col = col_idx

            dimensoes[sheet_name] = (last_row, last_col)
    finally:
        wb_read.close()

    # 2º passo: abrir normalmente e remover o excesso
    wb = openpyxl.load_workbook(file_path, keep_vba=keep_vba)
    for sheet_name, (last_row, last_col) in dimensoes.items():
        ws = wb[sheet_name]

        if last_row == 0 or last_col == 0:
            continue

        max_row = ws.max_row
        max_col = ws.max_column

        rows_removed = max(0, max_row - last_row)
        cols_removed = max(0, max_col - last_col)

        if cols_removed > 0:
            ws.delete_cols(last_col + 1, cols_removed)
        if rows_removed > 0:
            ws.delete_rows(last_row + 1, rows_removed)

        if rows_removed > 0 or cols_removed > 0:
            logger.info(f"[LIMPEZA] Sheet '{sheet_name}': {rows_removed} linhas e {cols_removed} colunas vazias removidas (dados até {last_row}x{last_col})")

    return wb


def _set_cell_value(ws, row, column, value):
    """
    Define o valor de uma célula, desviando para a célula superior-esquerda
    caso o destino seja uma MergedCell (somente leitura no openpyxl).
    """
    cell = ws.cell(row=row, column=column)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.row >= merged_range.min_row and cell.row <= merged_range.max_row \
                    and cell.column >= merged_range.min_col and cell.column <= merged_range.max_col:
                top_left = merged_range.start_cell
                ws.cell(row=top_left.row, column=top_left.column, value=value)
                logger.debug(f"[EDITAR_PLANILHA] Valor '{value}' escrito na célula mesclada {top_left.coordinate} (requisitado {cell.coordinate})")
                return
    else:
        cell.value = value


def _editar_sheet_sumario(wb, dados_modificados, data_competencia):
    if 'Sumario' not in wb.sheetnames:
        logger.warning("[EDITAR_PLANILHA] Sheet 'Sumario' não encontrada")
        return

    ws = wb['Sumario']
    logger.debug("[EDITAR_PLANILHA] Editando sheet 'Sumario'")

    if data_competencia:
        valor_competencia = data_competencia if isinstance(data_competencia, str) else data_competencia.strftime('%d/%m/%Y')
        _set_cell_value(ws, row=6, column=1, value=valor_competencia)
        logger.debug(f"[EDITAR_PLANILHA] Sumario - competência atualizada: {valor_competencia}")

    total_condos = len(dados_modificados.get('condominios', []))
    total_func = sum(len(c.get('funcionarios', [])) for c in dados_modificados.get('condominios', []))

    _set_cell_value(ws, row=8, column=2, value=total_condos)
    _set_cell_value(ws, row=9, column=2, value=total_func)

    logger.info(f"[EDITAR_PLANILHA] Sumario atualizado: {total_condos} condominios, {total_func} funcionarios")


def _editar_sheet_local_entrega(wb, dados_modificados):
    if 'Local de Entrega' not in wb.sheetnames:
        logger.warning("[EDITAR_PLANILHA] Sheet 'Local de Entrega' não encontrada")
        return

    ws = wb['Local de Entrega']
    logger.debug("[EDITAR_PLANILHA] Editando sheet 'Local de Entrega'")

    max_row = ws.max_row
    if max_row > 1:
        ws.delete_rows(2, max_row - 1)
        logger.debug(f"[EDITAR_PLANILHA] Local de Entrega - {max_row - 1} linhas antigas removidas")

    for row_idx, condo in enumerate(dados_modificados.get('condominios', []), start=2):
        cnpj = ''.join(filter(str.isdigit, str(condo.get('cnpj', ''))))
        _set_cell_value(ws, row=row_idx, column=1, value=cnpj)
        _set_cell_value(ws, row=row_idx, column=2, value=condo.get('nome', ''))
        _set_cell_value(ws, row=row_idx, column=3, value='AV')
        _set_cell_value(ws, row=row_idx, column=4, value=condo.get('rua', ''))
        _set_cell_value(ws, row=row_idx, column=5, value=condo.get('numero', ''))
        _set_cell_value(ws, row=row_idx, column=6, value=condo.get('complemento', ''))
        _set_cell_value(ws, row=row_idx, column=7, value=condo.get('bairro', ''))
        _set_cell_value(ws, row=row_idx, column=8, value=condo.get('cidade', ''))
        _set_cell_value(ws, row=row_idx, column=9, value=condo.get('estado', ''))
        _set_cell_value(ws, row=row_idx, column=10, value=condo.get('cep', ''))
        logger.debug(f"[EDITAR_PLANILHA] Local de Entrega - linha {row_idx}: cnpj={cnpj}, nome={condo.get('nome', '')}, cep={condo.get('cep', '')}")

    logger.info(f"[EDITAR_PLANILHA] Local de Entrega atualizado: {len(dados_modificados.get('condominios', []))} linhas")


def _editar_sheet_beneficiario(wb, dados_modificados):
    if 'Beneficiario' not in wb.sheetnames:
        logger.warning("[EDITAR_PLANILHA] Sheet 'Beneficiario' não encontrada")
        return

    ws = wb['Beneficiario']
    logger.debug("[EDITAR_PLANILHA] Editando sheet 'Beneficiario'")

    col_produtos = _ler_headers_produtos(ws)
    logger.debug(f"[EDITAR_PLANILHA] Headers de produtos mapeados: {col_produtos}")

    max_row = ws.max_row
    if max_row > 2:
        ws.delete_rows(3, max_row - 2)
        logger.debug(f"[EDITAR_PLANILHA] Beneficiario - {max_row - 2} linhas antigas removidas")

    row_num = 3
    for condo in dados_modificados.get('condominios', []):
        cnpj = ''.join(filter(str.isdigit, str(condo.get('cnpj', ''))))

        for func in condo.get('funcionarios', []):
            cpf = ''.join(filter(str.isdigit, str(func.get('cpf', '')))).zfill(11)
            data_nasc = func.get('data_nascimento', '')
            if data_nasc and isinstance(data_nasc, str):
                try:
                    dt = datetime.strptime(data_nasc, '%Y-%m-%d')
                    data_nasc = dt.strftime('%d%m%Y')
                except ValueError:
                    pass
            elif data_nasc and hasattr(data_nasc, 'strftime'):
                data_nasc = data_nasc.strftime('%d%m%Y')

            _set_cell_value(ws, row=row_num, column=1, value=cpf)
            _set_cell_value(ws, row=row_num, column=2, value=cnpj)
            _set_cell_value(ws, row=row_num, column=3, value='')
            _set_cell_value(ws, row=row_num, column=4, value=func.get('matricula', ''))
            _set_cell_value(ws, row=row_num, column=5, value=func.get('nome', ''))
            _set_cell_value(ws, row=row_num, column=6, value='')
            _set_cell_value(ws, row=row_num, column=7, value=data_nasc)
            _set_cell_value(ws, row=row_num, column=8, value=func.get('sexo', ''))
            _set_cell_value(ws, row=row_num, column=9, value='')

            logger.debug(f"[EDITAR_PLANILHA] Beneficiario - linha {row_num}: cpf={cpf}, cnpj={cnpj}, nome={func.get('nome', '')}, matricula={func.get('matricula', '')}, data_nasc={data_nasc}")

            for mov in func.get('movimentacoes', []):
                prod_nome = mov.get('produto', '')
                valor = mov.get('valor', 0)

                col_idx = _encontrar_coluna_produto(col_produtos, prod_nome)
                if col_idx:
                    _set_cell_value(ws, row=row_num, column=col_idx, value=float(valor))
                    logger.debug(f"[EDITAR_PLANILHA] Beneficiario - produto '{prod_nome}' na coluna {col_idx}, valor {valor}")
                else:
                    logger.warning(f"[EDITAR_PLANILHA] Beneficiario - produto '{prod_nome}' não encontrado nas colunas mapeadas")

            row_num += 1

    logger.info(f"[EDITAR_PLANILHA] Beneficiario atualizado: {row_num - 3} funcionarios")


def _ler_headers_produtos(ws):
    col_produtos = {}
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    if not header_row:
        logger.warning("[EDITAR_PLANILHA] Nenhum header de produto encontrado na linha 2 do sheet 'Beneficiario'")
        return col_produtos

    header_row = header_row[0]
    for col_idx, val in enumerate(header_row, start=1):
        if val is None:
            continue
        h = str(val).strip().split('\n')[0].strip()
        if not h:
            continue
        if h in COLUNAS_PRODUTO:
            col_produtos[col_idx] = h
            logger.debug(f"[EDITAR_PLANILHA] Header exato mapeado - coluna {col_idx}: {h}")
        else:
            for nome in COLUNAS_PRODUTO:
                if nome.lower() in h.lower() or h.lower() in nome.lower():
                    col_produtos[col_idx] = nome
                    logger.debug(f"[EDITAR_PLANILHA] Header aproximado mapeado - coluna {col_idx}: '{h}' -> '{nome}'")
                    break

    logger.info(f"[EDITAR_PLANILHA] Total de headers de produtos mapeados: {len(col_produtos)}")
    return col_produtos


def _encontrar_coluna_produto(col_produtos, nome_produto):
    if not nome_produto:
        logger.debug("[EDITAR_PLANILHA] Nome do produto vazio ao buscar coluna")
        return None

    for col_idx, nome in col_produtos.items():
        if nome == nome_produto:
            logger.debug(f"[EDITAR_PLANILHA] Produto '{nome_produto}' encontrado na coluna {col_idx} (correspondência exata)")
            return col_idx

    nome_lower = nome_produto.lower()
    for col_idx, nome in col_produtos.items():
        if nome.lower() in nome_lower or nome_lower in nome.lower():
            logger.debug(f"[EDITAR_PLANILHA] Produto '{nome_produto}' encontrado na coluna {col_idx} (correspondência parcial com '{nome}')")
            return col_idx

    logger.debug(f"[EDITAR_PLANILHA] Produto '{nome_produto}' não mapeado em nenhuma coluna")
    return None
