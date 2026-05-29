import logging
import re
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def parse_vt_excel(file_path, upload_id):
    """
    Parser específico para Vale Transporte no formato da planilha modelo.
    Usa openpyxl para ler diretamente as células sem depender de cabeçalhos do pandas.
    """
    try:
        # Carrega a planilha com openpyxl (mais controle sobre células mescladas)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        if 'USUARIOS' not in wb.sheetnames:
            return {"error": "Planilha não contém a aba 'USUARIOS'", "valido": False}
        
        sheet = wb['USUARIOS']
        
        # Encontra a linha do cabeçalho principal (linha com "CNPJ*")
        header_row_idx = None
        header_row2_idx = None
        
        for row_idx in range(1, min(sheet.max_row, 20)):
            cell_value = sheet.cell(row=row_idx, column=1).value
            if cell_value and str(cell_value).strip() == 'CNPJ*':
                header_row_idx = row_idx
                header_row2_idx = row_idx + 1  # Próxima linha tem os subcabeçalhos
                break
        
        if header_row_idx is None:
            return {"error": "Não foi possível localizar o cabeçalho 'CNPJ*' na planilha", "valido": False}
        
        # Mapeia as colunas de interesse baseado na posição
        # Baseado no formato da planilha:
        # Coluna A (1): CNPJ*
        # Coluna B (2): MATRÍCULA*
        # Coluna C (3): NOME COMPLETO*
        # Coluna D (4): EMAIL
        # Coluna E (5): CELULAR
        # Coluna F (6): ATIVO
        # Coluna G (7): ENDEREÇO*
        # Coluna H (8): CARGO
        # Coluna I (9): DEPARTAMENTO
        # Coluna J (10): DIAS TRABALHADOS*
        # Coluna K (11): CPF*
        # Coluna L (12): RG.
        # Coluna M (13): DG.
        # Coluna N (14): EST.RG
        # Coluna O (15): DATA DE NASCIMENTO
        # Coluna P (16): NOME DA MÃE
        # Coluna Q (17): LOGRADOURO
        # Coluna R (18): NÚMERO
        # Coluna S (19): COMPLEMENTO
        # Coluna T (20): BAIRRO
        # Coluna U (21): CEP
        # Coluna V (22): CIDADE
        # Coluna W (23): ESTADO
        # Coluna X (24): ITEM 1 (CÓD.)
        # Coluna Y (25): ITEM 1 (QTD.)
        # Coluna Z (26): ITEM 1 (DIAS.)
        # Coluna AA (27): ITEM 1 (VALOR)
        # E assim por diante...
        
        colunas = {
            'cnpj': 1,
            'matricula': 2,
            'nome_completo': 3,
            'email': 4,
            'celular': 5,
            'ativo': 6,
            'endereco': 7,
            'cargo': 8,
            'departamento': 9,
            'dias_trabalhados': 10,
            'cpf': 11,
            'rg': 12,
            'digito_rg': 13,
            'estado_rg': 14,
            'data_nascimento': 15,
            'nome_mae': 16,
            'logradouro': 17,
            'numero': 18,
            'complemento': 19,
            'bairro': 20,
            'cep': 21,
            'cidade': 22,
            'estado': 23,
        }
        
        # Processa as linhas de dados (começa após a linha do cabeçalho + 2)
        start_row = header_row_idx + 2
        
        erros = []
        dados_validados = []
        cpfs_validos = set()
        condominios_set = set()
        valor_total_vt = 0
        total_dias_trabalhados = 0
        total_registros = 0
        
        for row_idx in range(start_row, sheet.max_row + 1):
            # Pega o CNPJ da primeira coluna
            cnpj_cell = sheet.cell(row=row_idx, column=colunas['cnpj']).value
            if not cnpj_cell or str(cnpj_cell).strip() == '':
                continue  # Linha vazia, para de processar
            
            # Extrai dados básicos
            cnpj = str(cnpj_cell).strip()
            nome_funcionario = str(sheet.cell(row=row_idx, column=colunas['nome_completo']).value or '').strip()
            cpf_raw = str(sheet.cell(row=row_idx, column=colunas['cpf']).value or '').strip()
            cpf = re.sub(r'\D', '', cpf_raw)
            departamento = str(sheet.cell(row=row_idx, column=colunas['departamento']).value or '').strip()
            dias_trabalhados_raw = sheet.cell(row=row_idx, column=colunas['dias_trabalhados']).value
            
            # Converte dias trabalhados
            try:
                dias_trabalhados = int(float(dias_trabalhados_raw)) if dias_trabalhados_raw else 0
            except (ValueError, TypeError):
                dias_trabalhados = 0
            
            matricula = str(sheet.cell(row=row_idx, column=colunas['matricula']).value or '').strip()
            cargo = str(sheet.cell(row=row_idx, column=colunas['cargo']).value or '').strip()
            data_nascimento = sheet.cell(row=row_idx, column=colunas['data_nascimento']).value
            endereco = str(sheet.cell(row=row_idx, column=colunas['endereco']).value or '').strip()
            
            # Formata data de nascimento
            data_nascimento_str = ''
            if data_nascimento:
                if isinstance(data_nascimento, datetime):
                    data_nascimento_str = data_nascimento.strftime('%Y-%m-%d')
                else:
                    data_nascimento_str = str(data_nascimento)
            
            # Extrai CNPJ do condomínio do departamento (formato: "CNPJ - NOME")
            cnpj_condominio = cnpj
            nome_condominio = ''
            
            # Limpa o CNPJ (remove pontos, barras, hífens)
            cnpj_condominio_clean = re.sub(r'\D', '', cnpj)
            
            if departamento and '-' in departamento:
                parts = departamento.split('-', 1)
                cnpj_part = re.sub(r'\D', '', parts[0].strip())
                if len(cnpj_part) == 14:
                    cnpj_condominio_clean = cnpj_part
                    nome_condominio = parts[1].strip() if len(parts) > 1 else ''
            
            if not nome_condominio:
                nome_condominio = departamento or cnpj
            
            # Validações básicas
            if not cpf or len(cpf) != 11:
                erros.append({
                    "linha": row_idx,
                    "campo": "CPF",
                    "valor": cpf,
                    "mensagem": f"CPF inválido: {cpf}"
                })
                continue
            
            if not nome_funcionario:
                erros.append({
                    "linha": row_idx,
                    "campo": "NOME COMPLETO",
                    "valor": nome_funcionario,
                    "mensagem": "Nome do funcionário não informado"
                })
                continue
            
            if not cnpj_condominio_clean or len(cnpj_condominio_clean) != 14:
                erros.append({
                    "linha": row_idx,
                    "campo": "CNPJ",
                    "valor": cnpj_condominio_clean,
                    "mensagem": f"CNPJ do condomínio inválido (deve ter 14 dígitos): {cnpj_condominio_clean}"
                })
                continue
            
            # Extrai os itens (colunas 24 em diante, cada item ocupa 4 colunas)
            itens = []
            base_col = 24  # Coluna X (ITEM 1)
            
            for item_num in range(10):  # Máximo 10 itens
                col_cod = base_col + (item_num * 4)
                col_qtd = col_cod + 1
                col_dias = col_cod + 2
                col_valor = col_cod + 3
                
                # Verifica se a coluna existe
                if col_valor > sheet.max_column:
                    break
                
                codigo = sheet.cell(row=row_idx, column=col_cod).value
                qtd = sheet.cell(row=row_idx, column=col_qtd).value
                dias = sheet.cell(row=row_idx, column=col_dias).value
                valor = sheet.cell(row=row_idx, column=col_valor).value
                
                # Converte valores
                try:
                    qtd_num = int(float(qtd)) if qtd not in (None, '') else 1
                except (ValueError, TypeError):
                    qtd_num = 1
                
                try:
                    dias_num = int(float(dias)) if dias not in (None, '') else 0
                except (ValueError, TypeError):
                    dias_num = 0
                
                try:
                    # Remove R$ e converte
                    valor_str = str(valor).strip() if valor else ''
                    valor_str = valor_str.replace('R$', '').replace('.', '').replace(',', '.').strip()
                    valor_num = float(valor_str) if valor_str else 0
                except (ValueError, TypeError):
                    valor_num = 0
                
                codigo_str = str(codigo).strip() if codigo else ''
                
                if codigo_str and valor_num > 0:
                    itens.append({
                        'codigo': codigo_str,
                        'quantidade': qtd_num,
                        'dias': dias_num,
                        'valor': valor_num
                    })
            
            # Se não encontrou itens, tenta encontrar valor em qualquer coluna de valor
            if not itens:
                # Procura por qualquer valor não-zero nas colunas de valor dos itens
                for item_num in range(10):
                    col_valor = base_col + 3 + (item_num * 4)
                    if col_valor <= sheet.max_column:
                        valor = sheet.cell(row=row_idx, column=col_valor).value
                        try:
                            valor_str = str(valor).strip() if valor else ''
                            valor_str = valor_str.replace('R$', '').replace('.', '').replace(',', '.').strip()
                            valor_num = float(valor_str) if valor_str else 0
                            if valor_num > 0:
                                itens.append({
                                    'codigo': 'VT',
                                    'quantidade': 1,
                                    'dias': dias_trabalhados if dias_trabalhados > 0 else 22,
                                    'valor': valor_num
                                })
                                break
                        except:
                            pass
            
            if not itens:
                erros.append({
                    "linha": row_idx,
                    "campo": "ITENS",
                    "valor": "",
                    "mensagem": f"Nenhum item de vale transporte encontrado para {nome_funcionario}"
                })
                continue
            
            # Processa cada item do funcionário
            for item in itens:
                total_registros += 1
                valor_item = item['valor']
                dias_item = item['dias'] if item['dias'] > 0 else dias_trabalhados
                
                if valor_item <= 0:
                    erros.append({
                        "linha": row_idx,
                        "campo": "VALOR",
                        "valor": valor_item,
                        "mensagem": f"Valor do VT deve ser maior que 0 para {nome_funcionario}"
                    })
                    continue
                
                if dias_item <= 0:
                    dias_item = 22  # Valor padrão
                
                if dias_item > 31:
                    erros.append({
                        "linha": row_idx,
                        "campo": "DIAS",
                        "valor": dias_item,
                        "mensagem": f"Quantidade de dias não pode exceder 31 para {nome_funcionario}"
                    })
                    continue
                
                # Acumula totais
                cpfs_validos.add(cpf)
                if nome_condominio:
                    condominios_set.add(nome_condominio)
                
                valor_total_vt += valor_item
                total_dias_trabalhados += dias_item
                
                # Adiciona aos dados validados
                dados_validados.append({
                    "cnpj_condominio": cnpj_condominio_clean,
                    "nome_condominio": nome_condominio,
                    "cpf_funcionario": cpf,
                    "nome_funcionario": nome_funcionario,
                    "matricula": matricula,
                    "cargo": cargo,
                    "codigo_produto": item['codigo'],
                    "valor_vt": valor_item,
                    "quantidade_dias": dias_item,
                    "data_nascimento": data_nascimento_str,
                    "endereco": endereco,
                    "dias_trabalhados_mes": dias_trabalhados
                })
        
        # Determina se é válido
        valido = len([e for e in erros if "inválido" in e.get("mensagem", "") or "não informado" in e.get("mensagem", "")]) == 0 and len(dados_validados) > 0
        
        resultado = {
            "valido": valido,
            "mensagem_validacao": "Arquivo validado com sucesso" if valido else f"Encontrados {len(erros)} erros de validação",
            "total_registros": total_registros,
            "total_funcionarios": len(cpfs_validos),
            "total_condominios": len(condominios_set),
            "valor_total_vt": valor_total_vt,
            "total_dias_trabalhados": total_dias_trabalhados,
            "dados_validados": dados_validados,
            "linhas_com_erro": erros
        }
        
        logger.info(f"VT Parser resultado para upload {upload_id}: total_registros={total_registros}, valido={valido}")
        return resultado
        
    except Exception as e:
        logger.error(f"Erro no parser VT: {str(e)}", exc_info=True)
        return {"error": str(e), "valido": False}