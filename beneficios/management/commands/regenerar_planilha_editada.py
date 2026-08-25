"""
Regenera a planilha editada de importações a partir das movimentações
confirmadas no banco, usando o gerador corrigido (coluna por código do produto).

Uso:
    python manage.py regenerar_planilha_editada 233 247 --dry-run
    python manage.py regenerar_planilha_editada 233 247

Fluxo por importação: baixa a planilha ORIGINAL do S3 -> reconstrói
`dados_modificados` das MovimentacaoBeneficio -> editar_planilha_original →
valida o resultado (linhas com valor e soma == total confirmado) -> sobe no S3
com sufixo -REGERADO -> atualiza Importacao/FileUpload.arquivo_s3_editado.
O arquivo defeituoso anterior NÃO é apagado do S3 (auditoria).

Contexto: varredura de 25/08/2026 (docs/varredura-planilhas-editadas-20260825.csv).
"""
import os
import tempfile
from datetime import datetime
from decimal import Decimal
from urllib.parse import quote, unquote, urlparse

import boto3
from django.conf import settings
from django.core.management.base import BaseCommand
from django.db.models import Sum
from openpyxl import load_workbook

from beneficios.models import Importacao, MovimentacaoBeneficio
from upload.gerar_planilha_editada import editar_planilha_original

BUCKET = 'fedcorp-prod'


class Command(BaseCommand):
    help = "Regenera a planilha editada de importações a partir das movimentações confirmadas."

    def add_arguments(self, parser):
        parser.add_argument('importacao_ids', nargs='+', type=int)
        parser.add_argument('--dry-run', action='store_true', help="Gera e valida localmente, sem subir no S3 nem gravar no banco.")
        parser.add_argument('--out', default=None, help="Pasta para guardar cópia local dos arquivos gerados.")
        parser.add_argument('--insecure-ssl', action='store_true',
                            help="Desliga a verificação TLS do S3 (só para estações com proxy/CA corporativa; nunca no servidor).")

    def handle(self, *args, **options):
        s3 = boto3.client(
            's3',
            aws_access_key_id=getattr(settings, 'ACCESS_KEY_S3', ''),
            aws_secret_access_key=getattr(settings, 'SECRET_KEY_S3', ''),
            region_name='us-east-2',
            verify=not options['insecure_ssl'],
        )
        out_dir = options['out']
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)

        resumo = {'ok': 0, 'pulado': 0, 'erro': 0}
        for imp_id in options['importacao_ids']:
            try:
                resultado = self._processar(s3, imp_id, options['dry_run'], out_dir)
                resumo[resultado] += 1
            except Exception as e:
                resumo['erro'] += 1
                self.stdout.write(self.style.ERROR(f"[{imp_id}] ERRO: {e!r}"))

        self.stdout.write(self.style.SUCCESS(f"Resumo: {resumo}" + (" (dry-run)" if options['dry_run'] else "")))

    # ------------------------------------------------------------------
    def _processar(self, s3, imp_id, dry_run, out_dir):
        imp = Importacao.objects.select_related('administradora', 'file_upload').filter(id=imp_id).first()
        if not imp:
            self.stdout.write(self.style.WARNING(f"[{imp_id}] importação não encontrada — pulado"))
            return 'pulado'

        url_original = imp.arquivo_s3 or (imp.file_upload.arquivo_s3 if imp.file_upload else None)
        if not url_original:
            self.stdout.write(self.style.WARNING(f"[{imp_id}] sem planilha original no S3 — pulado"))
            return 'pulado'

        movs = (
            MovimentacaoBeneficio.objects
            .filter(importacao=imp)
            .select_related('empresa_cnpj', 'funcionario_cpf', 'produto_codigo')
            .order_by('empresa_cnpj__nome', 'funcionario_cpf__nome')
        )
        if not movs.exists():
            self.stdout.write(self.style.WARNING(f"[{imp_id}] sem movimentações no banco — pulado"))
            return 'pulado'

        dados, total_db, competencia = self._montar_dados(movs)

        # --- baixa original ---
        key_original = unquote(urlparse(url_original).path.lstrip('/'))
        ext = os.path.splitext(key_original)[1] or '.xlsm'
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp_path = tmp.name
        try:
            s3.download_file(BUCKET, key_original, tmp_path)
            planilha = editar_planilha_original(tmp_path, dados, competencia)
        finally:
            try:
                os.remove(tmp_path)
            except OSError:
                pass

        if not planilha:
            self.stdout.write(self.style.ERROR(f"[{imp_id}] gerador não produziu planilha (valores sem coluna?) — ver log"))
            return 'erro'

        # --- valida o gerado antes de subir ---
        linhas, com_valor, soma = self._inspecionar(planilha)
        n_func = sum(len(c['funcionarios']) for c in dados['condominios'])
        ok = linhas == n_func and com_valor == n_func and abs(soma - float(total_db)) < 0.05
        status_txt = (
            f"[{imp_id}] adm {imp.administradora_id} | {imp.status} | funcionarios={n_func} "
            f"linhas={linhas} com_valor={com_valor} soma={soma:.2f} confirmado={float(total_db):.2f}"
        )
        if not ok:
            self.stdout.write(self.style.ERROR(status_txt + " -> VALIDAÇÃO FALHOU, não enviado"))
            return 'erro'

        if out_dir:
            with open(os.path.join(out_dir, f'editada_{imp_id}_regerada{ext}'), 'wb') as f:
                planilha.seek(0)
                f.write(planilha.read())

        if dry_run:
            self.stdout.write(status_txt + " -> OK (dry-run, nada enviado)")
            return 'ok'

        # --- upload + referências ---
        adm_nome = " ".join(str(imp.administradora.razao_social if imp.administradora else 'SISTEMA').split()[:2])
        base_original = os.path.splitext(os.path.basename(key_original))[0]
        ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        key_editada = f"VR - DOCS/importacoes/editadas/{adm_nome}-EDITADO-{base_original}-{ts}-REGERADO{ext}"
        planilha.seek(0)
        s3.upload_fileobj(planilha, BUCKET, key_editada)
        url_editada = f"https://{BUCKET}.s3.us-east-2.amazonaws.com/{quote(key_editada)}"

        url_anterior = imp.arquivo_s3_editado
        Importacao.objects.filter(id=imp.id).update(arquivo_s3_editado=url_editada)
        if imp.file_upload_id:
            from upload.models import FileUpload
            FileUpload.objects.filter(id=imp.file_upload_id).update(arquivo_s3_editado=url_editada)

        self.stdout.write(self.style.SUCCESS(status_txt + f" -> REGERADA\n    nova: {url_editada}\n    anterior (mantida no S3): {url_anterior}"))
        return 'ok'

    # ------------------------------------------------------------------
    @staticmethod
    def _montar_dados(movs):
        condos = {}
        total = Decimal('0')
        competencia = None
        for m in movs:
            cond = m.empresa_cnpj
            c = condos.setdefault(cond.cnpj, {
                'cnpj': cond.cnpj, 'nome': cond.nome,
                'rua': cond.endereco or '', 'numero': cond.numero or '', 'complemento': cond.complemento or '',
                'bairro': cond.bairro or '', 'cidade': cond.cidade or '', 'estado': cond.estado or '', 'cep': cond.cep or '',
                'funcionarios': {},
            })
            func = m.funcionario_cpf
            f = c['funcionarios'].setdefault(func.cpf, {
                'cpf': func.cpf, 'nome': func.nome, 'matricula': func.matricula or '',
                'data_nascimento': func.data_nascimento.isoformat() if func.data_nascimento else '',
                'sexo': func.sexo or '', 'movimentacoes': [],
            })
            f['movimentacoes'].append({
                'produto': m.produto_codigo.nome,
                'codigo_produto': m.produto_codigo.codigo_produto,
                'valor': float(m.valor_beneficio or 0),
            })
            total += (m.valor_beneficio or 0)
            competencia = competencia or m.data_competencia
        dados = {'condominios': [dict(c, funcionarios=list(c['funcionarios'].values())) for c in condos.values()]}
        return dados, total, competencia

    @staticmethod
    def _inspecionar(planilha):
        planilha.seek(0)
        wb = load_workbook(planilha, read_only=True, data_only=True)
        ws = wb['Beneficiario']
        linhas = com_valor = 0
        soma = 0.0
        # Template atual: dados da linha 3; template antigo: da linha 2. Em vez
        # de posição fixa, reconhece linha de dados pelo CPF na coluna A.
        for r in ws.iter_rows(min_row=1, values_only=True):
            if not r or r[0] is None:
                continue
            cpf = ''.join(ch for ch in str(r[0]) if ch.isdigit())
            if len(cpf) < 11 or not str(r[0]).strip().replace('.', '').replace('-', '').isdigit():
                continue  # cabeçalho ou linha sem CPF
            linhas += 1
            vals = [v for i, v in enumerate(r[:30]) if i >= 9 and isinstance(v, (int, float)) and v]
            if vals:
                com_valor += 1
                soma += sum(vals)
        wb.close()
        planilha.seek(0)
        return linhas, com_valor, soma
