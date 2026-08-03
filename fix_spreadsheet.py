#!/usr/bin/env python3
"""
Script para corrigir a planilha VR - CONDOMINIOS - ATUALIZADA (2).xlsm
1. Remove 3 linhas sem benefício (CPFs específicos)
2. Adiciona condomínio CNPJ 62296207000102 na aba 'Local de Entrega'
"""
import openpyxl
import re
import shutil
import os

ARQUIVO_ORIGINAL = "/home/daniel/Área de trabalho/CODES/VR-BCK/VR - CONDOMINIOS - ATUALIZADA (2).xlsm"
ARQUIVO_BACKUP = ARQUIVO_ORIGINAL.replace('.xlsm', '_BACKUP.xlsm')
ARQUIVO_CORRIGIDO = ARQUIVO_ORIGINAL

CPFS_REMOVER = ['26275522879', '34679402857', '02146898402']

CNPJ_ADICIONAR = {
    'codigo': '62296207000102',
    'nome': 'EDIFICIO PATIO TROPICAL',
    'rua': 'AVENIDA HORACIO LAFER',
    'numero': '370',
    'complemento': '',
    'bairro': 'JARDIM TEGEREBA',
    'cidade': 'GUARUJA',
    'estado': 'SP',
    'cep': '11430000',
}


def normalizar_cnpj(val):
    if val is None:
        return ''
    return re.sub(r'\D', '', str(val)).zfill(14)[:14]


def normalizar_cpf(val):
    if val is None:
        return ''
    return re.sub(r'\D', '', str(val)).zfill(11)[:11]


def main():
    # 1. Criar backup
    if not os.path.exists(ARQUIVO_BACKUP):
        shutil.copy2(ARQUIVO_ORIGINAL, ARQUIVO_BACKUP)
        print(f"Backup criado: {ARQUIVO_BACKUP}")
    else:
        print(f"Backup já existe: {ARQUIVO_BACKUP}")

    # 2. Abrir planilha (preservando macros VBA)
    print(f"Abrindo planilha: {ARQUIVO_ORIGINAL}")
    wb = openpyxl.load_workbook(ARQUIVO_ORIGINAL, keep_vba=True)

    # =============================
    # 3. ABA BENEFICIÁRIO - Remover linhas sem benefício
    # =============================
    ws_ben = wb['Beneficiario']
    print(f"\n--- Aba Beneficiário ---")
    print(f"Total de linhas: {ws_ben.max_row}")

    # Mapear CPFs das linhas a remover (buscar na coluna A = índice 0)
    linhas_para_remover = []
    cpfs_encontrados = set()

    for row_idx in range(2, ws_ben.max_row + 1):  # Pular header (linha 1)
        cpf_cell = ws_ben.cell(row=row_idx, column=1).value
        cpf_normalizado = normalizar_cpf(cpf_cell)

        if cpf_normalizado in CPFS_REMOVER:
            nome = ws_ben.cell(row=row_idx, column=5).value or ''
            print(f"  Linha {row_idx}: CPF {cpf_normalizado} - {nome} -> REMOVER")
            linhas_para_remover.append(row_idx)
            cpfs_encontrados.add(cpf_normalizado)

    # Remover do final para baixo (para não deslocar índices)
    linhas_para_remover.sort(reverse=True)
    for row_idx in linhas_para_remover:
        ws_ben.delete_rows(row_idx, 1)

    print(f"\n  Linhas removidas: {len(linhas_para_remover)}")
    print(f"  CPFs encontrados: {cpfs_encontrados}")
    cpfs_nao_encontrados = set(CPFS_REMOVER) - cpfs_encontrados
    if cpfs_nao_encontrados:
        print(f"  CPFs NÃO encontrados na planilha: {cpfs_nao_encontrados}")

    # =============================
    # 4. ABA LOCAL DE ENTREGA - Adicionar condomínio ausente
    # =============================
    ws_locais = wb['Local de Entrega']
    print(f"\n--- Aba Local de Entrega ---")
    print(f"Total de linhas: {ws_locais.max_row}")

    # Verificar se o CNPJ já existe
    cnpj_existe = False
    for row_idx in range(2, ws_locais.max_row + 1):
        codigo_cell = ws_locais.cell(row=row_idx, column=1).value
        if normalizar_cnpj(codigo_cell) == CNPJ_ADICIONAR['codigo']:
            cnpj_existe = True
            print(f"  CNPJ {CNPJ_ADICIONAR['codigo']} já existe na linha {row_idx}")
            break

    if cnpj_existe:
        # Atualizar dados do condomínio existente
        ws_locais.cell(row=row_idx, column=4, value=CNPJ_ADICIONAR['rua'])
        ws_locais.cell(row=row_idx, column=5, value=CNPJ_ADICIONAR['numero'])
        ws_locais.cell(row=row_idx, column=6, value=CNPJ_ADICIONAR['complemento'])
        ws_locais.cell(row=row_idx, column=7, value=CNPJ_ADICIONAR['bairro'])
        ws_locais.cell(row=row_idx, column=8, value=CNPJ_ADICIONAR['cidade'])
        ws_locais.cell(row=row_idx, column=9, value=CNPJ_ADICIONAR['estado'])
        ws_locais.cell(row=row_idx, column=10, value=CNPJ_ADICIONAR['cep'])
        print(f"  CNPJ {CNPJ_ADICIONAR['codigo']} atualizado na linha {row_idx}")
    else:
        # Adicionar na próxima linha disponível
        nova_linha = ws_locais.max_row + 1
        ws_locais.cell(row=nova_linha, column=1, value=CNPJ_ADICIONAR['codigo'])
        ws_locais.cell(row=nova_linha, column=2, value=CNPJ_ADICIONAR['nome'])
        ws_locais.cell(row=nova_linha, column=3, value='')  # Matrícula
        ws_locais.cell(row=nova_linha, column=4, value=CNPJ_ADICIONAR['rua'])
        ws_locais.cell(row=nova_linha, column=5, value=CNPJ_ADICIONAR['numero'])
        ws_locais.cell(row=nova_linha, column=6, value=CNPJ_ADICIONAR['complemento'])
        ws_locais.cell(row=nova_linha, column=7, value=CNPJ_ADICIONAR['bairro'])
        ws_locais.cell(row=nova_linha, column=8, value=CNPJ_ADICIONAR['cidade'])
        ws_locais.cell(row=nova_linha, column=9, value=CNPJ_ADICIONAR['estado'])
        ws_locais.cell(row=nova_linha, column=10, value=CNPJ_ADICIONAR['cep'])
        print(f"  CNPJ {CNPJ_ADICIONAR['codigo']} adicionado na linha {nova_linha}")

    # =============================
    # 5. SALVAR
    # =============================
    print(f"\nSalvando planilha corrigida...")
    wb.save(ARQUIVO_CORRIGIDO)
    print(f"Planilha salva: {ARQUIVO_CORRIGIDO}")
    print("\n✅ Correção concluída!")


if __name__ == '__main__':
    main()
